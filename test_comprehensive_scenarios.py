"""
既存の全機能・全制約条件を対象に、極端なエッジケースを含む複数シナリオで
自動最適化・整合性検証を行う網羅的検証スクリプト。

test_shift_integrity.py が「デフォルトシナリオ1本」を対象にした基本検証で
あるのに対し、本スクリプトは以下4つの異なるシナリオそれぞれについて
ソルバーを実行し、各シナリオごとに重点チェック項目(希望休・有給確定の
反映精度、自動生成シフトの整合性、テーブル構造・Excel出力・手動編集の
堅牢性)を検証する。重複ロジックは test_shift_integrity.py の既存チェック
関数を単一の真実の情報源として再利用する。

シナリオ:
  A. 通常月シナリオ          - 標準的な希望休・有休データでの基本検証
  B. 希望休・有休集中シナリオ - 特定週末・特定店舗への希望休/有給確定の集中
  C. 特別休業日3日設定シナリオ - 盆・年末年始+3日相当の臨時休業を追加
  D. 未来月先行入力・月度切り替えシナリオ - 9→10→11→9→10月度を往復

実行方法:
    python test_comprehensive_scenarios.py
"""

from __future__ import annotations

import datetime as dt
import io
import os
import traceback

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import test_shift_integrity as tsi
import utils
from optimizer import solve_shift

RESULTS: list[tuple[str, str, bool, str]] = []  # (scenario, check name, passed, detail)


def record(scenario: str, name: str, issues: list[str]) -> bool:
    passed = len(issues) == 0
    detail = "OK" if passed else "; ".join(issues[:15]) + (f" ...(他{len(issues) - 15}件)" if len(issues) > 15 else "")
    RESULTS.append((scenario, name, passed, detail))
    return passed


def _req(name_to_id: dict, name: str, date: dt.date, kind: str) -> dict:
    return {"staff_id": name_to_id[name], "name": name, "date": date, "kind": kind}


# ---------------------------------------------------------------------------
# 共通チェック(4シナリオ共通で使い回す)。
# 【重点チェック項目1】希望休・有給確定の反映精度
# ---------------------------------------------------------------------------

def check_confirmed_leave_zero_attendance(shift_df: pd.DataFrame, requests_df: pd.DataFrame) -> list[str]:
    """「有給確定」が設定された日は、例外なく全スタッフが店舗出勤ゼロになっているか。"""
    issues = []
    if requests_df.empty:
        return issues
    confirmed = requests_df[requests_df["kind"] == "有給確定"]
    working_pairs = set(zip(shift_df["staff_id"], shift_df["date"])) if not shift_df.empty else set()
    for _, r in confirmed.iterrows():
        if (r["staff_id"], r["date"]) in working_pairs:
            issues.append(f"{r['name']} {r['date']}: 有給確定日にもかかわらず出勤している")
    return issues


def check_salaried_workdays_exact(
    shift_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
    requests_df: pd.DataFrame,
    base_holiday_quota: int,
) -> list[str]:
    """店長・正社員の出勤日数が「基準出勤日数 − 期間内の有給確定日数」に厳格一致しているか。

    基準出勤日数そのものは対象期間のカレンダーから動的に導出する(固定値を
    ハードコードしない)。有給確定を1件も持たないスタッフは「基準出勤日数と
    厳格一致」、有給確定をn件持つスタッフは「基準出勤日数-n日」を期待値とする。
    """
    issues = []
    expected_workdays = len(dates_df) - base_holiday_quota
    salaried = staff_df[staff_df["emp_type"].isin(["店長", "正社員"])]
    business_days = set(dates_df.loc[dates_df["is_business_day"], "date"])
    for _, row in salaried.iterrows():
        sid = row["staff_id"]
        cnt = len(shift_df[shift_df["staff_id"] == sid]) if not shift_df.empty else 0
        confirmed_cnt = 0
        if not requests_df.empty:
            confirmed_cnt = int(
                (
                    (requests_df["staff_id"] == sid)
                    & (requests_df["kind"] == "有給確定")
                    & (requests_df["date"].isin(business_days))
                ).sum()
            )
        expected = expected_workdays - confirmed_cnt
        if cnt != expected:
            issues.append(
                f"{row['name']}: 出勤{cnt}日"
                f"(期待値{expected}日 = 基準{expected_workdays}日-有給確定{confirmed_cnt}日)"
            )
    return issues


def check_individual_off_day_preserved(shift_df: pd.DataFrame, staff_df: pd.DataFrame, dates_df: pd.DataFrame) -> list[str]:
    """店長・正社員が営業日の中で最低1日は個別公休を確保できているか(ハード制約の健全性)。"""
    issues = []
    n_business_days = int(dates_df["is_business_day"].sum())
    salaried = staff_df[staff_df["emp_type"].isin(["店長", "正社員"])]
    for _, row in salaried.iterrows():
        sid = row["staff_id"]
        cnt = len(shift_df[shift_df["staff_id"] == sid]) if not shift_df.empty else 0
        if n_business_days >= 1 and cnt > n_business_days - 1:
            issues.append(f"{row['name']}: 出勤{cnt}日(営業日{n_business_days}日中、個別公休1日が確保されていない)")
    return issues


# ---------------------------------------------------------------------------
# 【重点チェック項目4】テーブル構造・UI連動・エラーログ
# ---------------------------------------------------------------------------

def check_manual_shift_wide_structure(shift_df: pd.DataFrame, dates_df: pd.DataFrame, staff_df: pd.DataFrame) -> list[str]:
    """店舗別日別シフト表(手動編集用ワイド形式)の行構成・空欄処理・変換の健全性を検証する。

    行数・枠番号は utils.STORE_MAX_HEADCOUNT(実際の店舗別上限人数の単一の真実の
    情報源)から動的に導出する。7店舗中、上限3名(徳重店・名古屋中川店・
    天白植田店)が3店舗、上限2名(稲沢店・大治店・新蟹江店・極楽店)が4店舗
    という実際の店舗容量制約を反映するため、行数は3*3+4*2=17行になる。
    """
    issues = []
    wide = utils.build_manual_shift_wide(shift_df, dates_df)

    expected_rows = sum(utils.STORE_MAX_HEADCOUNT.get(s, 2) for s in utils.STORES)
    if len(wide) != expected_rows:
        issues.append(f"手動編集ワイド表の行数が想定と異なる(想定{expected_rows}行、実際{len(wide)}行)")

    for store in utils.STORES:
        sub = wide[wide["店舗"] == store]
        expected_slots = list(range(1, utils.STORE_MAX_HEADCOUNT.get(store, 2) + 1))
        actual_slots = sorted(sub["枠"].tolist())
        if actual_slots != expected_slots:
            issues.append(f"{store}: 枠番号が想定と異なる(想定{expected_slots}、実際{actual_slots})")

    date_labels = utils.manual_shift_date_labels(dates_df)
    for label in date_labels:
        if (wide[label] == utils.BLANK_LABEL).any():
            issues.append(f"{label}: 生データ(空文字列であるべき)にBLANK_LABELがそのまま含まれている")

    long_df = utils.manual_shift_wide_to_long(wide, dates_df, staff_df)
    if len(long_df) != len(shift_df):
        issues.append(f"wide→long変換で件数が一致しない(元{len(shift_df)}件、変換後{len(long_df)}件)")

    # 画面表示(空文字列→「（空白）」に変換した状態)からの逆変換でも、
    # 件数が失われず・BLANK_LABELという名前のスタッフが誤って生成されないか。
    display_wide = wide.copy()
    for label in date_labels:
        display_wide[label] = display_wide[label].replace("", utils.BLANK_LABEL)
    long_from_display = utils.manual_shift_wide_to_long(display_wide, dates_df, staff_df)
    if len(long_from_display) != len(shift_df):
        issues.append(
            f"「（空白）」表示状態からのwide→long変換で件数が一致しない"
            f"(元{len(shift_df)}件、変換後{len(long_from_display)}件)"
        )
    if not long_from_display.empty and (long_from_display["name"] == utils.BLANK_LABEL).any():
        issues.append("「（空白）」ラベルがスタッフ名として誤って長形式に変換されている")

    return issues


def check_manual_alert_extreme_edge_cases(staff_df: pd.DataFrame, dates_df: pd.DataFrame) -> list[str]:
    """手動編集アラート再判定(utils.check_manual_shift_alerts)の極端な入力への耐性。

    KeyError・IndexError・TypeError等の未処理例外が一切発生しないことを検証する。
    """
    issues = []
    business_day0 = dates_df.loc[dates_df["is_business_day"], "date"].iloc[0]
    sid0, name0, emp0 = staff_df["staff_id"].iloc[0], staff_df["name"].iloc[0], staff_df["emp_type"].iloc[0]
    sid1, name1, emp1 = staff_df["staff_id"].iloc[1], staff_df["name"].iloc[1], staff_df["emp_type"].iloc[1]

    edge_cases = [
        ("完全に空のshift_df", pd.DataFrame(columns=["date", "store", "staff_id", "name", "emp_type"])),
        (
            "存在しないスタッフIDを含む1行",
            pd.DataFrame([{"date": business_day0, "store": utils.STORES[0], "staff_id": "NOPE", "name": "架空太郎", "emp_type": "パート"}]),
        ),
        (
            "存在しない店舗名を含む1行",
            pd.DataFrame([{"date": business_day0, "store": "架空店", "staff_id": sid0, "name": name0, "emp_type": emp0}]),
        ),
        (
            "同一スタッフが同日に2店舗重複配置",
            pd.DataFrame(
                [
                    {"date": business_day0, "store": utils.STORES[0], "staff_id": sid0, "name": name0, "emp_type": emp0},
                    {"date": business_day0, "store": utils.STORES[1], "staff_id": sid0, "name": name0, "emp_type": emp0},
                ]
            ),
        ),
        (
            "全営業日・全店舗が空欄(誰も出勤していない)",
            pd.DataFrame(columns=["date", "store", "staff_id", "name", "emp_type"]),
        ),
        (
            "同一店舗に同一スタッフの重複+別スタッフの正常配置が混在",
            pd.DataFrame(
                [
                    {"date": business_day0, "store": utils.STORES[0], "staff_id": sid0, "name": name0, "emp_type": emp0},
                    {"date": business_day0, "store": utils.STORES[0], "staff_id": sid0, "name": name0, "emp_type": emp0},
                    {"date": business_day0, "store": utils.STORES[0], "staff_id": sid1, "name": name1, "emp_type": emp1},
                ]
            ),
        ),
    ]
    for label, df in edge_cases:
        try:
            alerts = utils.check_manual_shift_alerts(df, staff_df, dates_df)
            if not isinstance(alerts, dict) or not all(k in alerts for k in ("shortages", "skill_issues", "duplicates")):
                issues.append(f"{label}: 戻り値の構造が想定外({alerts})")
        except Exception as e:  # noqa: BLE001 - 「例外が出ないこと」自体を検証する
            issues.append(f"{label}: 例外が発生しクラッシュした({type(e).__name__}: {e})")
    return issues


def check_excel_export_styles(
    shift_df: pd.DataFrame,
    dates_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    requests_df: pd.DataFrame,
) -> list[str]:
    """Excel出力(フォント9pt・行高39pt・列幅5.4・全店舗4行構成)がスタイルエラーなく
    反映されているか。"""
    try:
        xlsx_bytes = utils.build_export_workbook(shift_df, dates_df, staff_df, requests_df)
    except Exception as e:
        return [f"Excel生成中に例外が発生した({type(e).__name__}: {e})"]

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes))
    except Exception as e:
        return [f"生成されたExcelファイルの読み込みに失敗した({type(e).__name__}: {e})"]

    issues = []
    n_date_cols = len(dates_df)
    STORE_SLOT_ROWS = 4  # 全7店舗を一律4行(1〜4枠)で出力する仕様(utils.pyと同じ値)
    header_and_store_rows_end = 1 + STORE_SLOT_ROWS * len(utils.STORES)  # ヘッダー1行+7店舗×4行=29行目

    for sheet_name in ["店舗別日別シフト表", "スタッフ別出勤一覧表"]:
        if sheet_name not in wb.sheetnames:
            issues.append(f"シート「{sheet_name}」が存在しない")
            continue
        ws = wb[sheet_name]

        header_cell = ws.cell(row=1, column=1)
        if header_cell.font.size != 9:
            issues.append(f"{sheet_name}: 1行目フォントサイズが9ptでない(実際{header_cell.font.size})")
        if not header_cell.font.bold:
            issues.append(f"{sheet_name}: 1行目が太字でない(既存スタイルが崩れている)")

        if sheet_name == "店舗別日別シフト表":
            h1 = ws.row_dimensions[1].height
            if h1 != 39.0:
                issues.append(f"{sheet_name}: 1行目の行高が39pt(52px相当)でない(実際{h1})")
            for r in range(2, header_and_store_rows_end + 1):
                h = ws.row_dimensions[r].height
                if h != 28.5:
                    issues.append(f"{sheet_name}: {r}行目の行高が28.5pt(38px相当)でない(実際{h})")
        else:
            for r in range(1, 19):
                h = ws.row_dimensions[r].height
                if h != 39.0:
                    issues.append(f"{sheet_name}: {r}行目の行高が39pt(52px相当)でない(実際{h})")

        if sheet_name == "店舗別日別シフト表":
            for col_idx in range(2, 34):
                w = ws.column_dimensions[get_column_letter(col_idx)].width
                if w != 5.4:
                    issues.append(f"{sheet_name}: {get_column_letter(col_idx)}列の幅が5.4(43px相当)でない(実際{w})")

            # 全7店舗が一律4行(28行)構成になっているかを検証する。
            expected_store_at_row = {}
            row_cursor = 2
            for store in utils.STORES:
                expected_store_at_row[row_cursor] = store
                row_cursor += STORE_SLOT_ROWS
            for row, expected_store in expected_store_at_row.items():
                actual_store = ws.cell(row=row, column=1).value
                if actual_store != expected_store:
                    issues.append(
                        f"{sheet_name}: {row}行目の店舗名が想定と異なる(想定={expected_store}, 実際={actual_store})"
                    )
            last_row = header_and_store_rows_end
            if row_cursor - 1 != last_row:
                issues.append(
                    f"{sheet_name}: 7店舗×4行の合計行数が想定と異なる(想定終端={last_row}行目, 実際={row_cursor - 1}行目)"
                )
        else:
            for col_idx in range(3, 3 + n_date_cols):
                w = ws.column_dimensions[get_column_letter(col_idx)].width
                if w != 5.4:
                    issues.append(f"{sheet_name}: {get_column_letter(col_idx)}列の幅が5.4(43px相当)でない(実際{w})")

    return issues


def check_shortage_attribution_accuracy(shortages: list[dict]) -> list[str]:
    """負荷・エッジケースシナリオ(希望休・有給確定の意図的な集中)では、店舗の
    必要人員に対する不足が現実的に発生し得る(optimizer.py が「不足/超過スラック
    変数＋重いペナルティ」で意図的にソフト化している領域であり、7店舗×2名の
    週末必要人員=14枠に対し社員・嘱託の総数もちょうど14名という無遊び構成の
    ため、週末に1名でも休むとどこかで必ず不足が生じる)。そのため不足の発生
    そのものはNGとしない。

    代わりに、不足が発生した場合は必ず「原因候補」が特定され(=原因不明・
    説明不能な不足が発生していないか)、管理者が画面上で理由を追えることを
    検証する。
    """
    issues = []
    for s in shortages:
        cause = s.get("原因候補", "")
        if not cause or cause.startswith("(希望休以外の要因"):
            issues.append(f"{s['date']} {s['store']}: 不足の原因候補が特定できない({s['内容']}, 不足{s['不足数']})")
    return issues


def check_day_range_shortfall_reported_accurately(result, staff_df: pd.DataFrame) -> list[str]:
    """嘱託・パートの契約勤務日数レンジは、optimizer.py の設計方針として明示的に
    「不足/超過スラック変数＋重いペナルティ」のソフト制約として実装されている
    (特別休業日の集中等でカレンダーの営業日数が大きく減った月は、実績データに
    基づく固定目標(例: 前田=11日固定)を満たせないことがあり得る、という
    ドキュメント化された設計)。そのためレンジ未達そのものはNGとしない。

    代わりに、レンジ未達/超過が発生した場合、その旨が必ずスタッフサマリー
    (staff_summary_df の「勤務日数レンジ(嘱託/パート)」列)へ正しく反映され、
    管理者が画面上で気づける状態になっているか(=ソフト制約からの逸脱が
    サイレントに握りつぶされていないか)を検証する。
    """
    issues = []
    for _, row in staff_df.iterrows():
        if row["emp_type"] not in ("嘱託", "パート"):
            continue
        sid = row["staff_id"]
        cnt = len(result.shift_df[result.shift_df["staff_id"] == sid]) if not result.shift_df.empty else 0
        if row["emp_type"] == "嘱託":
            lo, hi = row["min_workdays"], row["max_workdays"]
        else:
            lo, hi = utils.PART_ROLE_WORKDAY_RANGE.get(row["part_role"], (None, None))
        if lo is None or pd.isna(lo):
            continue
        summary_row = result.staff_summary_df[result.staff_summary_df["staff_id"] == sid]
        if summary_row.empty:
            issues.append(f"{row['name']}: スタッフサマリーに該当行が存在しない")
            continue
        note = str(summary_row.iloc[0]["勤務日数レンジ(嘱託/パート)"])
        if cnt < lo and "下限" not in note:
            issues.append(f"{row['name']}: 出勤{cnt}日で下限{int(lo)}日を下回っているが、サマリーに不足が反映されていない(表示: {note})")
        elif hi is not None and pd.notna(hi) and cnt > hi and "上限" not in note:
            issues.append(f"{row['name']}: 出勤{cnt}日で上限{int(hi)}日を上回っているが、サマリーに超過が反映されていない(表示: {note})")
    return issues


# ---------------------------------------------------------------------------
# ソルバー実行後の共通検証セット(重点チェック項目1〜3)
# ---------------------------------------------------------------------------

def run_common_post_solve_checks(
    scenario: str,
    dates_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    requests_df: pd.DataFrame,
    result,
    base_holiday_quota: int,
    strict_day_ranges: bool = True,
    strict_store_headcount: bool = True,
) -> None:
    record(scenario, "有給確定日の出勤ゼロ", check_confirmed_leave_zero_attendance(result.shift_df, requests_df))
    record(
        scenario,
        "店長・正社員の出勤日数(基準日数-有給確定日数に厳格一致)",
        check_salaried_workdays_exact(result.shift_df, staff_df, dates_df, requests_df, base_holiday_quota),
    )
    record(scenario, "個別公休1日の確保", check_individual_off_day_preserved(result.shift_df, staff_df, dates_df))

    real_shortages, skill_issues, duplicates = tsi.check_store_requirements(result, staff_df, dates_df)
    record(scenario, "測定・加工スキル保持者の同席(全店舗・全営業日)", skill_issues)
    record(scenario, "同日重複配置の非存在", duplicates)
    if strict_store_headcount:
        record(scenario, "店舗成立要件(不足許容日を除く)", real_shortages)
        record(scenario, "名古屋中川店 土日祝 社員2名以上", tsi.check_nakagawa_weekend(result, staff_df, dates_df))
    else:
        # 希望休・有給確定を意図的に集中させた負荷シナリオでは、週末必要人員
        # (7店舗×2名=14枠)と社員・嘱託の総数(14名)がちょうど釣り合っている
        # ため、週末に1名でも休むとどこかで必ず不足が生じる(optimizer.pyが
        # 意図的にソフト制約として設計している領域)。不足の発生そのものは
        # NGとせず、原因が正しく特定・表示されているかを検証する。
        record(
            scenario,
            "店舗成立要件(不足発生時は原因候補の特定精度を検証)",
            check_shortage_attribution_accuracy(result.shortages),
        )
    record(scenario, "個別スタッフの勤務制限(若松/若林/中村)", tsi.check_individual_restrictions(result, staff_df, dates_df))
    record(
        scenario,
        "稲沢店における若松優先配置(竹内=稲沢+若松=他店の禁止)",
        tsi.check_wakamatsu_takeuchi_inazawa_priority(result, staff_df),
    )
    if strict_day_ranges:
        record(scenario, "嘱託・パートの契約勤務日数レンジ", tsi.check_day_ranges(result, staff_df))
    else:
        # 特別休業日の集中等でカレンダーの営業日数が大きく減った月は、実績データに
        # 基づく固定目標(例: 前田=11日固定)を満たせないことがある(optimizer.pyが
        # 意図的にソフト制約として設計している領域)。厳格一致ではなく、逸脱が
        # 正しく可視化されているかを検証する。
        record(
            scenario,
            "嘱託・パートの契約勤務日数レンジ(未達時はサマリーへの反映精度を検証)",
            check_day_range_shortfall_reported_accurately(result, staff_df),
        )
    record(
        scenario,
        "有休代替候補の妥当性(正社員除外/上限未到達)",
        tsi.check_leave_availability_integrity(result, staff_df, dates_df, requests_df),
    )


# ---------------------------------------------------------------------------
# シナリオA: 通常月シナリオ
# ---------------------------------------------------------------------------

def scenario_a_normal_month() -> None:
    scenario = "A:通常月"
    year, month = 2026, 9
    dates = utils.get_period_dates(year, month)
    dates_df = utils.classify_days(dates)
    staff_df = utils.default_staff_df()
    business_days = list(dates_df.loc[dates_df["is_business_day"], "date"])
    # 「通常月」シナリオは典型的な希望休・有休の入り方を想定するため、店長・
    # 正社員の有給確定は平日に限定する(土日祝は週末必要人員14枠(7店舗×2名)
    # に対し社員・嘱託の総数もちょうど14名という無遊び構成のため、土日祝の
    # 有給確定はどのスタッフでも構造的に不足を生みやすい=シナリオBで別途
    # 意図的に検証する負荷エッジケースであり、こちらでは対象外とする)。
    weekday_business_days = [d for d in business_days if d.weekday() not in (5, 6)]
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))

    requests_rows = [
        _req(name_to_id, "生駒", weekday_business_days[3], "有給確定"),
        _req(name_to_id, "加藤", weekday_business_days[10], "希望休"),
        _req(name_to_id, "尾澤", weekday_business_days[6], "絶対休"),
        _req(name_to_id, "柴田", weekday_business_days[12], "希望休"),
    ]
    probe_df = pd.DataFrame(requests_rows, columns=["staff_id", "name", "date", "kind"])
    avail = utils.compute_paid_leave_availability(dates_df, staff_df, probe_df)
    slot_days = avail.loc[avail["paid_leave_slots"] > 0, "date"].tolist()
    if slot_days:
        requests_rows.append(_req(name_to_id, "田中", slot_days[0], "有給申請"))

    requests_df = pd.DataFrame(requests_rows, columns=["staff_id", "name", "date", "kind"])
    closed_days = int((~dates_df["is_business_day"]).sum())
    base_holiday_quota = closed_days + 1

    result = solve_shift(dates_df, staff_df, requests_df, base_holiday_quota=base_holiday_quota, time_limit_sec=60)
    if not record(scenario, "求解可能性(FEASIBLE/OPTIMAL)", [] if result.is_feasible else [f"ステータス={result.status_name}"]):
        return

    run_common_post_solve_checks(scenario, dates_df, staff_df, requests_df, result, base_holiday_quota)

    record(
        scenario,
        "店舗別日別シフト表の行構成・空欄処理・wide/long変換の健全性",
        check_manual_shift_wide_structure(result.shift_df, dates_df, staff_df),
    )
    record(scenario, "手動編集アラート再判定の極端な入力への耐性", check_manual_alert_extreme_edge_cases(staff_df, dates_df))
    record(
        scenario,
        "Excel出力(フォント9pt/行高39pt/列幅5.4)のスタイル整合性",
        check_excel_export_styles(result.shift_df, dates_df, staff_df, requests_df),
    )


# ---------------------------------------------------------------------------
# シナリオB: 希望休・有休集中シナリオ(負荷・エッジケース)
# ---------------------------------------------------------------------------

def scenario_b_concentrated_leave() -> None:
    scenario = "B:希望休有休集中"
    year, month = 2026, 9
    dates = utils.get_period_dates(year, month)
    dates_df = utils.classify_days(dates)
    staff_df = utils.default_staff_df()
    business_days = list(dates_df.loc[dates_df["is_business_day"], "date"])
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))

    weekend_days = [d for d in business_days if d.weekday() in (5, 6)]
    holiday_only_days = list(
        dates_df.loc[
            dates_df["is_business_day"] & dates_df["is_weekend_or_holiday"] & (~dates_df["date"].isin(weekend_days)),
            "date",
        ]
    )
    if len(weekend_days) < 2:
        record(scenario, "検証用の土日データ確保", ["対象期間に十分な土日が存在しない"])
        return
    w1, w2 = weekend_days[0], weekend_days[1]

    # 中川店・徳重店・稲沢店それぞれの主力スタッフの希望休・有給確定が
    # 同一週末に重複・集中するケース(=各店の主戦力が同時に抜ける負荷シナリオ)。
    requests_rows = [
        _req(name_to_id, "内田", w1, "有給確定"),   # 名古屋中川店長
        _req(name_to_id, "真田", w1, "希望休"),      # 正社員(全店ヘルプ可)
        _req(name_to_id, "加藤", w1, "有給確定"),    # 徳重店長
        _req(name_to_id, "生駒", w2, "有給確定"),    # 稲沢店長
        _req(name_to_id, "山岡", w2, "希望休"),      # 嘱託(稲沢主所属)
    ]
    if holiday_only_days:
        h1 = holiday_only_days[0]
        requests_rows.append(_req(name_to_id, "小林", h1, "有給確定"))  # 新蟹江店長
        requests_rows.append(_req(name_to_id, "竹内", h1, "有給申請"))  # 嘱託(全店ヘルプ可)

    requests_df = pd.DataFrame(requests_rows, columns=["staff_id", "name", "date", "kind"])
    closed_days = int((~dates_df["is_business_day"]).sum())
    base_holiday_quota = closed_days + 1

    result = solve_shift(dates_df, staff_df, requests_df, base_holiday_quota=base_holiday_quota, time_limit_sec=90)
    if not record(scenario, "求解可能性(FEASIBLE/OPTIMAL・負荷集中下でもクラッシュしない)", [] if result.is_feasible else [f"ステータス={result.status_name}"]):
        return

    run_common_post_solve_checks(
        scenario, dates_df, staff_df, requests_df, result, base_holiday_quota, strict_store_headcount=False
    )


# ---------------------------------------------------------------------------
# シナリオC: 特別休業日3日設定シナリオ(盆・年末年始+3日相当)
# ---------------------------------------------------------------------------

def scenario_c_special_closures() -> None:
    scenario = "C:特別休業日3日"
    year, month = 2026, 9
    dates = utils.get_period_dates(year, month)
    dates_df_baseline = utils.classify_days(dates)
    staff_df = utils.default_staff_df()
    requests_df = utils.default_requests_df()

    business_days_baseline = list(dates_df_baseline.loc[dates_df_baseline["is_business_day"], "date"])
    special_closure_dates = business_days_baseline[2:5]  # 3日分の臨時休業を追加

    dates_df = utils.classify_days(dates, special_closure_dates=special_closure_dates)

    closed_days_baseline = int((~dates_df_baseline["is_business_day"]).sum())
    closed_days = int((~dates_df["is_business_day"]).sum())
    issues_count = []
    if closed_days != closed_days_baseline + 3:
        issues_count.append(f"定休日数が+3日になっていない(通常{closed_days_baseline}日→特別休業込み{closed_days}日)")
    record(scenario, "特別休業日3日分が定休日数へ正しく加算されているか", issues_count)

    base_holiday_quota_baseline = closed_days_baseline + 1
    base_holiday_quota = closed_days + 1
    expected_workdays_baseline = len(dates_df_baseline) - base_holiday_quota_baseline
    expected_workdays = len(dates_df) - base_holiday_quota

    issues_formula = []
    if expected_workdays != expected_workdays_baseline - 3:
        issues_formula.append(
            f"出勤日数の減算が特別休業3日分になっていない(通常{expected_workdays_baseline}日→{expected_workdays}日)"
        )
    if expected_workdays_baseline == 21 and expected_workdays != 18:
        issues_formula.append(
            f"通常月21日を基準とした場合の期待値18日と不一致(実際{expected_workdays}日)"
        )
    record(scenario, "特別休業日3日による正社員出勤日数の減算式(21日→18日相当)", issues_formula)

    # 個別公休1日はどちらの状態でも変わらず1日のみ(=特別休業日が個別公休枠を
    # 侵食も上乗せもしない)であることを確認する。
    target_off_baseline = max(0, base_holiday_quota_baseline - closed_days_baseline)
    target_off = max(0, base_holiday_quota - closed_days)
    issues_target_off = []
    if target_off_baseline != 1 or target_off != 1:
        issues_target_off.append(f"個別公休1日の想定が崩れている(通常{target_off_baseline}日、特別休業込み{target_off}日)")
    record(scenario, "個別公休1日が特別休業日の有無に関わらず維持されているか", issues_target_off)

    result = solve_shift(dates_df, staff_df, requests_df, base_holiday_quota=base_holiday_quota, time_limit_sec=60)
    if not record(scenario, "求解可能性(FEASIBLE/OPTIMAL)", [] if result.is_feasible else [f"ステータス={result.status_name}"]):
        return

    issues_closure_off = []
    if not result.shift_df.empty:
        for d in special_closure_dates:
            if not result.shift_df[result.shift_df["date"] == d].empty:
                issues_closure_off.append(f"{d}: 特別休業日にもかかわらず出勤者が存在する")
    record(scenario, "特別休業日は全店舗・全スタッフが休みになっているか", issues_closure_off)

    run_common_post_solve_checks(
        scenario, dates_df, staff_df, requests_df, result, base_holiday_quota, strict_day_ranges=False
    )


# ---------------------------------------------------------------------------
# シナリオD: 未来月先行入力・月度切り替えシナリオ
# ---------------------------------------------------------------------------

def scenario_d_future_month_switching() -> None:
    scenario = "D:月度切り替え"
    from streamlit.testing.v1 import AppTest

    months = [(2026, 9), (2026, 10), (2026, 11)]
    app_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.py")
    target_paths = [utils.kyuka_log_path_for(y, m) for y, m in months] + [
        utils.LATEST_SHIFT_PATH,
        utils.LATEST_SHIFT_META_PATH,
    ]
    backups: dict[str, bytes] = {}
    for p in target_paths:
        if os.path.exists(p):
            with open(p, "rb") as f:
                backups[p] = f.read()
            os.remove(p)
        lock_p = p + ".lock"
        if os.path.exists(lock_p):
            os.remove(lock_p)

    issues: list[str] = []
    try:
        at = AppTest.from_file(app_path, default_timeout=180)
        at.run()
        if list(at.exception):
            record(scenario, "初回描画で例外が発生しないこと", [str(at.exception)])
            return

        at.sidebar.button[0].click().run()
        if list(at.exception):
            issues.append(f"9月度・初回最適化で例外: {at.exception}")

        # 9月度に希望休を1件追加(個別申請フォームと同じ経路=ログへの直接追記)。
        utils.append_kyuka_request("生駒", dt.date(2026, 9, 20), "希望休", utils.kyuka_log_path_for(2026, 9))
        at.run()
        if list(at.exception):
            issues.append(f"9月度再描画で例外: {at.exception}")
        if len(at.session_state["requests_df"]) != 1 or at.session_state["requests_df"].iloc[0]["name"] != "生駒":
            issues.append("9月度: ログへの追記が次の再描画で反映されない")

        # 10月度へ切り替え、先行入力する。
        at.sidebar.selectbox[0].select(10).run()
        if list(at.exception):
            issues.append(f"10月度への切り替えで例外: {at.exception}")
        if len(at.session_state["requests_df"]) != 0:
            issues.append("10月度切り替え直後、希望休データが空になっていない(月度分離の不備)")
        if at.session_state["solve_result"] is not None:
            issues.append("10月度切り替え後もsolve_resultが古い月度のまま残っている")

        utils.append_kyuka_request("辻本", dt.date(2026, 10, 5), "有給確定", utils.kyuka_log_path_for(2026, 10))
        at.run()
        if list(at.exception):
            issues.append(f"10月度再描画で例外: {at.exception}")
        if len(at.session_state["requests_df"]) != 1 or at.session_state["requests_df"].iloc[0]["name"] != "辻本":
            issues.append("10月度: ログへの追記が次の再描画で反映されない")

        # 11月度へ切り替え(まだ何も入力していない、未来月の先行入力を想定)。
        at.sidebar.selectbox[0].select(11).run()
        if list(at.exception):
            issues.append(f"11月度への切り替えで例外: {at.exception}")
        if len(at.session_state["requests_df"]) != 0:
            issues.append("11月度: 他月度のデータが混入している(月度分離の不備)")

        utils.append_kyuka_request("尾澤", dt.date(2026, 11, 2), "希望休", utils.kyuka_log_path_for(2026, 11))
        at.run()
        if list(at.exception):
            issues.append(f"11月度再描画で例外: {at.exception}")
        if len(at.session_state["requests_df"]) != 1 or at.session_state["requests_df"].iloc[0]["name"] != "尾澤":
            issues.append("11月度: ログへの追記が次の再描画で反映されない")

        # 9月度・10月度へ順に戻り、それぞれのデータが往復後も破損していないか確認する。
        at.sidebar.selectbox[0].select(9).run()
        if list(at.exception):
            issues.append(f"9月度への復帰で例外: {at.exception}")
        restored9 = at.session_state["requests_df"]
        if len(restored9) != 1 or restored9.iloc[0]["name"] != "生駒":
            issues.append(f"9月度復帰後のデータが破損している(件数={len(restored9)})")

        at.sidebar.selectbox[0].select(10).run()
        if list(at.exception):
            issues.append(f"10月度への再切り替えで例外: {at.exception}")
        restored10 = at.session_state["requests_df"]
        if len(restored10) != 1 or restored10.iloc[0]["name"] != "辻本":
            issues.append(f"10月度のデータが破損している(件数={len(restored10)})")

        tab3 = at.tabs[2]
        if not any("最適化を実行" in i.value for i in tab3.info):
            issues.append("未計算月度(10月)でタブ3に案内メッセージが表示されない")

        record(scenario, "9→10→11→9→10月度往復でのクラッシュ非発生・データ分離・破損なし", issues)
    finally:
        for p in target_paths:
            if p in backups:
                os.makedirs(os.path.dirname(p), exist_ok=True)
                with open(p, "wb") as f:
                    f.write(backups[p])
            elif os.path.exists(p):
                os.remove(p)
            lock_p = p + ".lock"
            if os.path.exists(lock_p):
                os.remove(lock_p)


# ---------------------------------------------------------------------------
# メイン実行
# ---------------------------------------------------------------------------

def _run_scenario_safely(scenario_id: str, description: str, fn) -> None:
    print(f"\n--- シナリオ{scenario_id}: {description} を実行中... ---")
    try:
        fn()
    except Exception as e:  # noqa: BLE001 - 未処理例外の検出自体がこのスクリプトの目的
        tb = traceback.format_exc()
        record(scenario_id, "シナリオ実行中の未処理例外", [f"{type(e).__name__}: {e}"])
        print(tb)


def main() -> int:
    print("=" * 78)
    print("網羅的シナリオ検証(エッジケース含む) - test_comprehensive_scenarios.py")
    print("=" * 78)

    _run_scenario_safely("A:通常月", "通常月シナリオ", scenario_a_normal_month)
    _run_scenario_safely("B:希望休有休集中", "希望休・有休集中シナリオ(負荷・エッジケース)", scenario_b_concentrated_leave)
    _run_scenario_safely("C:特別休業日3日", "特別休業日3日設定シナリオ", scenario_c_special_closures)
    _run_scenario_safely("D:月度切り替え", "未来月先行入力・月度切り替えシナリオ", scenario_d_future_month_switching)

    print()
    print("-" * 78)
    print("検証結果一覧")
    print("-" * 78)
    n_pass = 0
    n_fail = 0
    current_scenario = None
    for scenario, name, passed, detail in RESULTS:
        if scenario != current_scenario:
            print(f"\n[{scenario}]")
            current_scenario = scenario
        mark = "✅ PASS" if passed else "❌ NG"
        if passed:
            n_pass += 1
        else:
            n_fail += 1
        print(f"  [{mark}] {name}")
        if not passed:
            print(f"           └ {detail}")

    print()
    print("-" * 78)
    print(f"合計: {len(RESULTS)}件 / PASS {n_pass}件 / NG {n_fail}件")
    print("=" * 78)
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
