"""
眼鏡店7店舗 シフト最適化アプリ - 共通ユーティリティ

このモジュールが担当する範囲:
  - シフト対象期間（毎月16日〜翌月15日）のカレンダー生成と営業区分判定
  - 店舗マスタ・スタッフマスタのデフォルトデータ
  - 有給休暇の「取得可能日・人数枠」の自動算出と案内文生成
  - Excel（.xlsx）エクスポート（店舗別日別シフト表 / スタッフ別出勤一覧表）

「社長」は現場出勤ゼロが絶対要件のため、スタッフマスタ・シフト変数のいずれにも
一切登場しない（=そもそもデータとして持たない）。
"""

from __future__ import annotations

import calendar
import colorsys
import csv
import datetime as dt
import io
import json
import os
import re
import time
import unicodedata
from dataclasses import dataclass, field
from typing import Optional

import jpholiday
import pandas as pd

# ---------------------------------------------------------------------------
# 店舗マスタ
# ---------------------------------------------------------------------------

STORES = [
    "稲沢店",
    "大治店",
    "名古屋中川店",
    "新蟹江店",
    "徳重店",
    "極楽店",
    "天白植田店",
]

# パートの役割 -> 勤務可能店舗（絶対制限）
PART_ROLE_ALLOWED_STORES = {
    "A": ["大治店", "名古屋中川店"],
    "B": ["徳重店"],
    "C": ["徳重店"],
    "D": ["天白植田店"],
    "E": ["大治店"],
}

# 「社員1名＋パート1名」体制が正式に許可されている店舗と、対応するパート役割
# （徳重店は特殊ルールのため別扱い）
COMBO_STORE_PART_ROLES = {
    "大治店": ["A", "E"],
    "名古屋中川店": ["A"],
    "天白植田店": ["D"],
}

TOKUSHIGE_STORE = "徳重店"
TOKUSHIGE_PART_ROLES = ["B", "C"]

GENERAL_STORES = [s for s in STORES if s not in COMBO_STORE_PART_ROLES and s != TOKUSHIGE_STORE]

# 「社員」として店舗の必要体制カウントに算入される区分（パートを除く）
EMPLOYEE_TYPES = ["店長", "正社員", "嘱託"]

# 店舗ごとの1日あたり最大人数キャップ(optimizer.pyのハード制約と一致させる)。
# Excel出力のレイアウト(店舗ごとの表示行数)にも利用する。
STORE_MAX_HEADCOUNT = {
    store: (3 if store in (TOKUSHIGE_STORE, "名古屋中川店", "天白植田店") else 2) for store in STORES
}

# パートの役割ごとの、期間内(16日〜翌月15日)勤務日数の許容範囲 (min, max)
# リソースを使い切らせすぎず、かつ最低限は稼働させるためのハード制約(重ペナルティ)。
PART_ROLE_WORKDAY_RANGE = {
    "A": (10, 13),  # 尾澤（大治店/名古屋中川店）
    "B": (12, 13),  # 不動野（徳重店）
    "C": (11, 11),  # 前田（徳重店）：実績データに準拠し11日固定
    "D": (11, 14),  # 柴田（天白植田店）
    "E": (12, 13),  # 野道（大治店）
}

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

NORMAL_HOURS = "10:00〜19:00"
SHORT_HOURS = "10:00〜17:00"


# ---------------------------------------------------------------------------
# カレンダー生成
# ---------------------------------------------------------------------------

def get_period_dates(year: int, month: int) -> list[dt.date]:
    """対象年月の「16日始まり〜翌月15日締め」の日付リストを返す。"""
    start = dt.date(year, month, 16)
    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1
    end = dt.date(next_year, next_month, 15)
    days = (end - start).days
    return [start + dt.timedelta(days=i) for i in range(days + 1)]


def last_tuesday_of_month(year: int, month: int) -> dt.date:
    """指定した暦月における「最終火曜日」を返す。"""
    last_day = calendar.monthrange(year, month)[1]
    d = dt.date(year, month, last_day)
    while d.weekday() != 1:  # Tuesday == 1
        d -= dt.timedelta(days=1)
    return d


def is_weekend_or_holiday(d: dt.date) -> bool:
    """土曜・日曜、または日本の祝日かどうかを判定する（土日祝の配置制約用）。"""
    return d.weekday() in (5, 6) or jpholiday.is_holiday(d)


def classify_days(
    dates: list[dt.date],
    special_closure_dates: list[dt.date] | None = None,
) -> pd.DataFrame:
    """各日の曜日・定休日/特別営業/通常営業/特別休業日の区分を判定する。

    ルール:
      - 毎週水曜日: 定休日
      - 火曜日: 「その暦月の最終火曜日」のみ特別営業(10-17時)、それ以外は定休日
      - 上記以外: 通常営業(10-19時)
      - special_closure_dates で指定された日(お盆・年末年始等の任意の全店一斉休業日)は、
        通常なら営業日となる日を強制的に「特別休業日」として休業扱いにする
        (=公休日数の計算式に「特別休業日数」として別枠で加算される)。
        既に定休日の日を指定しても二重カウントはしない。
    """
    special_set = set(special_closure_dates or [])

    # 期間内に登場しうる暦月それぞれの最終火曜日を事前計算
    months = sorted({(d.year, d.month) for d in dates})
    last_tue = {(y, m): last_tuesday_of_month(y, m) for y, m in months}

    rows = []
    for d in dates:
        weekday = d.weekday()  # Mon=0 ... Sun=6
        if weekday == 2:
            day_type, hours, note = "定休日", "-", "毎週水曜定休"
        elif weekday == 1:
            if d == last_tue[(d.year, d.month)]:
                day_type, hours, note = "特別営業(短縮)", SHORT_HOURS, "月内最終火曜日"
            else:
                day_type, hours, note = "定休日", "-", "火曜定休(最終火曜以外)"
        else:
            day_type, hours, note = "通常営業", NORMAL_HOURS, ""

        is_special_closure = False
        if d in special_set and day_type != "定休日":
            day_type, hours, note = "特別休業日", "-", "特別休業(お盆・年末年始等)"
            is_special_closure = True

        holiday_name = jpholiday.is_holiday_name(d) or ""
        rows.append(
            {
                "date": d,
                "weekday_jp": WEEKDAY_JP[weekday],
                "day_type": day_type,
                "hours": hours,
                "note": note,
                "is_closed": day_type in ("定休日", "特別休業日"),
                "is_business_day": day_type not in ("定休日", "特別休業日"),
                "is_weekend_or_holiday": is_weekend_or_holiday(d),
                "holiday_name": holiday_name,
                "is_special_closure": is_special_closure,
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# スタッフマスタ（デフォルト値・すぐに動作検証できるように用意）
# ---------------------------------------------------------------------------

def default_staff_df() -> pd.DataFrame:
    """店長7名＋正社員4名＋嘱託3名＋パート5名 = 19名のデフォルトマスタを生成する。

    社長は仕様上シフト要員から完全除外のため、このマスタには一切登場しない。
    店長は各自の店舗に完全固定（他店舗ヘルプ不可＝ハード制約）で、
    店長以外の正社員4名と嘱託3名のみが全店舗へのヘルプ出勤を許可される。
    """
    rows = []

    # 店長7名：各店舗に完全固定、他店ヘルプ不可、測定・加工スキル保有
    tencho_spec = [
        ("生駒", "稲沢店"),
        ("辻本", "大治店"),
        ("内田", "名古屋中川店"),
        ("小林", "新蟹江店"),
        ("加藤", "徳重店"),
        ("田中", "極楽店"),
        ("長瀬", "天白植田店"),
    ]
    for i, (name, home) in enumerate(tencho_spec):
        rows.append(
            dict(
                staff_id=f"T{i+1:02d}",
                name=name,
                emp_type="店長",
                home_store=home,
                allowed_stores=[home],
                can_help=False,
                custom_allowed_stores=None,
                part_role=None,
                has_skill=True,
                min_workdays=None,
            )
        )

    # 正社員4名（店長以外）：真田は全店舗ヘルプ可能。
    # 若林は勤務可能店舗が「極楽店(主所属)/徳重店/天白植田店」の3店舗のみに限定
    # (稲沢店・大治店・名古屋中川店・新蟹江店への配置は完全禁止)。
    # 中村は稲沢店への配置を禁止(勤務可能店舗から除外)。
    # 若松は勤務可能店舗が「新蟹江店/大治店/稲沢店/名古屋中川店」に限定され、
    # かつ測定・加工スキルを保有しない(=同席するスタッフに要スキル保有者、詳細は
    # 店舗必要体制のスキル要件が既に一般解として保証する)。
    seishain_spec = [
        ("真田", "名古屋中川店", True, None),
        ("若林", "極楽店", True, ["極楽店", "徳重店", "天白植田店"]),
        ("中村", "徳重店", True, [s for s in STORES if s != "稲沢店"]),
        ("若松", "新蟹江店", False, ["新蟹江店", "大治店", "稲沢店", "名古屋中川店"]),
    ]
    for i, (name, home, has_skill, custom_stores) in enumerate(seishain_spec):
        rows.append(
            dict(
                staff_id=f"S{i+1:02d}",
                name=name,
                emp_type="正社員",
                home_store=home,
                allowed_stores=list(custom_stores) if custom_stores else list(STORES),
                can_help=True,
                custom_allowed_stores=list(custom_stores) if custom_stores else None,
                part_role=None,
                has_skill=has_skill,
                min_workdays=None,
            )
        )

    # 嘱託3名：期間内勤務日数の範囲[min,max]あり。竹内は全7店舗にヘルプ出勤可能。
    shokutaku_spec = [
        ("吉田", "天白植田店", True, 16, 18, None),
        ("竹内", "天白植田店", True, 18, 20, None),
        ("山岡", "稲沢店", True, 18, 20, None),
    ]
    for i, (name, home, has_skill, min_days, max_days, custom_stores) in enumerate(shokutaku_spec):
        rows.append(
            dict(
                staff_id=f"K{i+1:02d}",
                name=name,
                emp_type="嘱託",
                home_store=home,
                allowed_stores=list(custom_stores) if custom_stores else list(STORES),
                can_help=True,
                custom_allowed_stores=list(custom_stores) if custom_stores else None,
                part_role=None,
                has_skill=has_skill,
                min_workdays=min_days,
                max_workdays=max_days,
            )
        )

    # パート5名：勤務可能店舗が絶対制限
    part_spec = [
        ("尾澤", "A"),
        ("不動野", "B"),
        ("前田", "C"),
        ("柴田", "D"),
        ("野道", "E"),
    ]
    for i, (name, role) in enumerate(part_spec):
        allowed = PART_ROLE_ALLOWED_STORES[role]
        rows.append(
            dict(
                staff_id=f"P{i+1:02d}",
                name=name,
                emp_type="パート",
                home_store=allowed[0],
                allowed_stores=list(allowed),
                can_help=False,
                custom_allowed_stores=None,
                part_role=role,
                has_skill=False,
                min_workdays=None,
            )
        )

    df = pd.DataFrame(rows)
    # 個別公休(残り1日)の希望日。既定は未指定(None)=AIが自動で最適配分する。
    # 実際に適用対象となるのは emp_type が「店長」「正社員」の11名のみ。
    df["preferred_off_date"] = None

    # 土日祝は特定の店舗への配置を禁止するスタッフ(ハード制約)。
    #   若松: 検査技能不足のため土日祝の新蟹江店・名古屋中川店配置を禁止(平日のみ)。
    #   野道: 大治店の土日祝出勤を禁止(平日のみ稼働)。
    df["weekend_holiday_forbidden_stores"] = [[] for _ in range(len(df))]
    df.loc[df["name"] == "若松", "weekend_holiday_forbidden_stores"] = df.loc[
        df["name"] == "若松", "weekend_holiday_forbidden_stores"
    ].apply(lambda _: ["新蟹江店", "名古屋中川店"])
    df.loc[df["name"] == "野道", "weekend_holiday_forbidden_stores"] = df.loc[
        df["name"] == "野道", "weekend_holiday_forbidden_stores"
    ].apply(lambda _: ["大治店"])

    # 出勤日にできる限りこの店舗へ配属してほしい、というソフトな優先店舗。
    #   吉田: 天白植田店を優先配属(他店ヘルプはやむを得ない場合のみ軽微なペナルティ)。
    #   山岡: 主所属の稲沢店を最優先配属(竹内・若松等のヘルプより優先させる)。
    df["preferred_store"] = None
    df.loc[df["name"] == "吉田", "preferred_store"] = "天白植田店"
    df.loc[df["name"] == "山岡", "preferred_store"] = "稲沢店"

    # 土曜・日曜(祝日は含まない)に限り特定の店舗への配置を禁止するスタッフ(ハード制約)。
    #   若松: 土日は大治店を優先配置とするため、残る稲沢店を禁止(=土日は大治店固定)。
    #        (新蟹江店は weekend_holiday_forbidden_stores で既に禁止済み)
    df["saturday_sunday_forbidden_stores"] = [[] for _ in range(len(df))]
    df.loc[df["name"] == "若松", "saturday_sunday_forbidden_stores"] = df.loc[
        df["name"] == "若松", "saturday_sunday_forbidden_stores"
    ].apply(lambda _: ["稲沢店"])

    return df


# 期間内に最低1回は特定の2名を同一店舗で同日勤務させる、という組み合わせ要件。
# (不足時はソルバーを落とさず重いペナルティで警告するソフト化ハード制約)
REQUIRED_PAIR_WORKDAYS = [
    {"names": ("辻本", "尾澤"), "store": "大治店", "min_occurrences": 1},
]

# 特定日・特定店舗の配置を手動で確定させる(ハード制約)。指定した氏名を必ず出勤
# させ、それ以外のスタッフは同日同店舗への配置を禁止する(=ロースターを完全固定)。
# 必要人数に満たない指定の場合は、その分だけ不足スラックが立ち、不足アラートに
# 反映される(ソルバーは落ちない)。対象期間に該当日が含まれない場合は無視される。
MANUAL_STORE_ASSIGNMENTS = [
    {
        "date": dt.date(2026, 10, 10),
        "store": "徳重店",
        "names": ["中村", "山岡"],
        "exclude_others": False,  # 必要人数(2名)は満たすため、パートの追加配置は妨げない
        "note": "土日祝の社員2名要件を充足(中村+山岡を確定配置)",
    },
    {
        "date": dt.date(2026, 10, 10),
        "store": "稲沢店",
        "names": ["生駒"],
        "exclude_others": True,  # 2人目を無理に埋めず1名不足を意図的に許容する
        "note": "生駒1名のみ配置。2人目は無理に埋めず1名不足を許容する。",
    },
]


def derive_allowed_stores(row) -> list[str]:
    """emp_type・部署固定ルール・個別の勤務可能店舗制限から勤務可能店舗を導出する。

    優先順位:
      1. 店長: 自店舗固定(ハード制約)。他の設定に関わらず home_store 1店舗のみ。
      2. パート: 役割ごとの絶対制限(勤務可能店舗の絶対制限)。
      3. custom_allowed_stores が設定されている場合(例: 若松・竹内): その店舗
         リストに限定(ハード制約)。他店舗ヘルプ許可フラグより優先される。
      4. それ以外(正社員・嘱託の標準ケース): 他店舗ヘルプ許可フラグに従い、
         全店舗 or 主所属店舗のみ。
    """

    def _get(key, default=None):
        if isinstance(row, dict):
            return row.get(key, default)
        return getattr(row, key, default)

    emp_type = _get("emp_type")
    home_store = _get("home_store")
    if emp_type == "店長":
        return [home_store]
    if emp_type == "パート":
        part_role = _get("part_role")
        return list(PART_ROLE_ALLOWED_STORES.get(part_role, [home_store]))
    custom_stores = _get("custom_allowed_stores")
    if isinstance(custom_stores, list) and len(custom_stores) > 0:
        return list(custom_stores)
    can_help = _get("can_help", True)
    return list(STORES) if can_help else [home_store]


def default_requests_df() -> pd.DataFrame:
    """希望休・有給申請の入力テーブルの空雛形。"""
    return pd.DataFrame(
        {
            "staff_id": pd.Series(dtype="str"),
            "name": pd.Series(dtype="str"),
            "date": pd.Series(dtype="object"),
            "kind": pd.Series(dtype="str"),  # 希望休 / 絶対休 / 有給申請
        }
    )


REQUEST_KINDS = ["希望休", "絶対休", "有給申請"]
HARD_OFF_KINDS = {"絶対休", "有給申請"}
SOFT_OFF_KINDS = {"希望休"}

# 旧表記からのデータ移行用エイリアス。過去に保存されたログ/CSVには、
#   (1) 「有休」表記(現行の「有給」表記へ改称する前)
#   (2) 「有給確定」区分(現行では廃止され「有給申請」に統合済み)
# のいずれかがそのまま残っている可能性があるため、読み込み時に必ずこの表を
# 通して現行の正規表記(「有給申請」)へ正規化する(=保存済みの絶対制約が、
# 表記変更・区分統合だけでハード制約から外れてしまう事故を防ぐ)。
_LEGACY_KIND_ALIASES = {
    "有休申請": "有給申請",
    "有休確定": "有給申請",  # 旧表記 かつ 廃止区分(二重の移行)
    "有給確定": "有給申請",  # 区分統合(「有給確定」は廃止し「有給申請」に一本化)
}


def _normalize_kind_value(kind: str) -> str:
    """kind文字列を、旧表記(有休)・廃止区分(有給確定)を含めて現行の正規表記
    (有給申請)へ正規化する。"""
    return _LEGACY_KIND_ALIASES.get(kind, kind)


# ---------------------------------------------------------------------------
# 希望休・有休入力テーブルのローカル永続化
# ---------------------------------------------------------------------------

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
SAVED_KYUKA_PATH = os.path.join(DATA_DIR, "saved_kyuka.csv")  # 旧・月度非分離のデフォルトパス(後方互換用)


def saved_kyuka_path_for(year: int, month: int) -> str:
    """月度別(シフト期間開始年月)の希望休・有休データの保存パスを返す。

    未来の月度(2〜3ヶ月先等)の希望休を先行入力・保存できるよう、月度ごとに
    完全に独立したファイルで管理する。
    """
    return os.path.join(DATA_DIR, f"saved_kyuka_{int(year)}_{int(month):02d}.csv")


def save_requests_to_disk(requests_df: pd.DataFrame, path: str = SAVED_KYUKA_PATH) -> None:
    """希望休・有休入力テーブルをローカルCSVへ即時保存する。

    画面での手入力・編集、CSV一括インポートのいずれの変更後にも呼び出すことで、
    アプリの再起動や別ブラウザ・別端末からのアクセス時にも内容が復元される。
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    out = requests_df.copy()
    if not out.empty:
        out["date"] = out["date"].apply(lambda d: d.isoformat() if isinstance(d, dt.date) else d)
    out.reindex(columns=["staff_id", "name", "date", "kind"]).to_csv(path, index=False, encoding="utf-8-sig")


def load_requests_from_disk(path: str = SAVED_KYUKA_PATH) -> pd.DataFrame:
    """ローカル保存済みの希望休・有休データを読み込む。存在しなければ空の雛形を返す。"""
    if not os.path.exists(path):
        return default_requests_df()
    try:
        df = pd.read_csv(path, dtype={"staff_id": str, "name": str, "kind": str}, encoding="utf-8-sig")
    except Exception:
        return default_requests_df()
    if df.empty or "date" not in df.columns:
        return default_requests_df()
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df = df.dropna(subset=["date"])
    df = df.reindex(columns=["staff_id", "name", "date", "kind"]).reset_index(drop=True)
    if "kind" in df.columns:
        df["kind"] = df["kind"].apply(lambda k: _normalize_kind_value(k) if isinstance(k, str) else k)
    return df


def clear_saved_requests(path: str = SAVED_KYUKA_PATH) -> None:
    """ローカル保存済みの希望休・有休データを削除する(リセット機能用)。"""
    if os.path.exists(path):
        os.remove(path)


# ---------------------------------------------------------------------------
# 希望休・有休の「個別申請・追記型ログ」管理
#
# 旧方式(save_requests_to_disk)は、画面上の全スタッフ分の入力テーブルを毎回
# まるごと1つのCSVへ上書き保存していた。このため、Aさんの端末が古いスナップ
# ショットを保持したまま(=Bさんが追加した申請をまだ知らない状態のまま)何らか
# の理由で再保存されると、Bさんの申請ごと消えてしまう競合が起こり得た。
#
# 新方式では、申請1件ごとに「誰が・いつ・何を申請したか」を1行としてログCSV
# (data/kyuka_requests_{year}_{month:02d}.csv)へ追記するのみとし、既存行の
# 読み込み→書き換え→全体保存は一切行わない。これにより、複数ブラウザ/端末から
# 同時に送信されても互いの行を上書き・消去することがなくなる(=追記は原理的に
# 衝突しない)。取消も「取消」種別の行を追記する論理削除として扱う。
# 現在の有効な申請一覧は、(スタッフ名, 日付)ごとにログの最新行を採用すること
# で都度再構築する(=イベントソーシング/追記ログの考え方)。
# ---------------------------------------------------------------------------

KYUKA_LOG_COLUMNS = ["staff_name", "date", "request_type", "updated_at"]
CANCELLED_REQUEST_TYPE = "取消"  # ログ上の取消(論理削除)マーカー


def kyuka_log_path_for(year: int, month: int) -> str:
    """月度別の希望休・有休「追記型ログ」の保存パスを返す。"""
    return os.path.join(DATA_DIR, f"kyuka_requests_{int(year)}_{int(month):02d}.csv")


def _kyuka_lock_path(path: str) -> str:
    return path + ".lock"


def _acquire_kyuka_lock(path: str, timeout_sec: float = 5.0, poll_interval: float = 0.02) -> None:
    """簡易な排他ロック(ロックファイル方式)を取得する。

    追記自体はOS上ほぼアトミックだが、ヘッダー行の初回書き込みと本体行の
    追記を1つの操作として直列化するため、念のためロックで保護する。
    タイムアウトした場合は、異常終了で残った古いロックとみなして強制解放し、
    保存処理自体が止まってしまわないようにする(可用性を優先)。
    """
    lock_path = _kyuka_lock_path(path)
    deadline = time.time() + timeout_sec
    while True:
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            return
        except FileExistsError:
            if time.time() > deadline:
                try:
                    os.remove(lock_path)
                except OSError:
                    pass
                continue
            time.sleep(poll_interval)


def _release_kyuka_lock(path: str) -> None:
    try:
        os.remove(_kyuka_lock_path(path))
    except OSError:
        pass


def append_kyuka_request(staff_name: str, date: dt.date, request_type: str, path: str) -> None:
    """個別の希望休・有給申請(または取消)を1件、ログCSVへ追記する。

    既存行の読み込み・書き換えを一切行わないため、他のスタッフが同時に別の
    申請を送信していても、互いのデータを消去することがない。
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    _acquire_kyuka_lock(path)
    try:
        file_is_new = not os.path.exists(path) or os.path.getsize(path) == 0
        with open(path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if file_is_new:
                writer.writerow(KYUKA_LOG_COLUMNS)
            writer.writerow(
                [
                    staff_name,
                    date.isoformat() if isinstance(date, dt.date) else str(date),
                    request_type,
                    dt.datetime.now().isoformat(timespec="seconds"),
                ]
            )
    finally:
        _release_kyuka_lock(path)


def load_kyuka_log(path: str) -> pd.DataFrame:
    """ログCSVの全行(履歴・取消行を含む)をそのまま読み込む。"""
    if not os.path.exists(path):
        return pd.DataFrame(columns=KYUKA_LOG_COLUMNS)
    try:
        df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    except Exception:
        return pd.DataFrame(columns=KYUKA_LOG_COLUMNS)
    if df.empty:
        return pd.DataFrame(columns=KYUKA_LOG_COLUMNS)
    return df.reindex(columns=KYUKA_LOG_COLUMNS)


def compute_current_requests_from_log(log_df: pd.DataFrame, staff_df: pd.DataFrame) -> pd.DataFrame:
    """追記型ログから、現在有効な希望休・有休の一覧(requests_df形式)を再構築する。

    同一の(スタッフ名, 日付)にログ行が複数ある場合は、ログの追記順で最後の行
    (=最新の申請内容)のみを採用する。最新行が「取消」だった場合は、その
    組み合わせを結果から除外する(論理削除)。
    """
    if log_df.empty:
        return default_requests_df()

    df = log_df.dropna(subset=["staff_name", "date"]).copy()
    if df.empty:
        return default_requests_df()
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    df = df.dropna(subset=["date"])
    if df.empty:
        return default_requests_df()
    # 過去に旧表記(有休申請/有休確定)で追記されたログ行も、新表記(有給)へ
    # 正規化してから採用する(表記変更だけでハード制約から外れないようにする)。
    df["request_type"] = df["request_type"].apply(lambda k: _normalize_kind_value(k) if isinstance(k, str) else k)

    # groupby(...).last() はグループ内の最終出現行(=ログ上で最も新しい申請)を採用する。
    latest = df.groupby(["staff_name", "date"], as_index=False, sort=False).last()
    latest = latest[latest["request_type"] != CANCELLED_REQUEST_TYPE]
    if latest.empty:
        return default_requests_df()

    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))
    latest = latest.copy()
    latest["staff_id"] = latest["staff_name"].map(name_to_id)
    latest = latest.dropna(subset=["staff_id"])
    latest = latest.rename(columns={"staff_name": "name", "request_type": "kind"})
    return latest.reindex(columns=["staff_id", "name", "date", "kind"]).reset_index(drop=True)


def load_current_requests(year: int, month: int, staff_df: pd.DataFrame) -> pd.DataFrame:
    """指定月度の「現在有効な」希望休・有休一覧を、追記型ログから再構築して返す。

    新方式のログファイルがまだ存在しない場合は、旧方式(全体上書き保存)の
    保存済みファイルが残っていればログへ1回だけ移行し(データを失わないため)、
    以降はログを正としてこの関数から状態を再構築する。
    """
    log_path = kyuka_log_path_for(year, month)
    if not os.path.exists(log_path):
        legacy_path = saved_kyuka_path_for(year, month)
        if os.path.exists(legacy_path):
            legacy_df = load_requests_from_disk(path=legacy_path)
            for _, r in legacy_df.iterrows():
                append_kyuka_request(r["name"], r["date"], r["kind"], log_path)

    log_df = load_kyuka_log(log_path)
    return compute_current_requests_from_log(log_df, staff_df)


def sync_admin_requests_edit(current_df: pd.DataFrame, edited_df: pd.DataFrame, path: str) -> None:
    """管理者によるマトリクス表の一括編集を、追記型ログへの差分追記に変換して保存する。

    `current_df`(編集前の集計状態)と`edited_df`(編集後の状態)を比較し、
    実際に値が変化した(スタッフ名, 日付)の組み合わせについてのみ新しいログ行
    (追加/変更は新種別、削除は「取消」)を追記する。変化のない申請には一切
    触れないため、この編集と同時に他のスタッフが送信した個別申請を巻き込んで
    消してしまうことがない。

    重要: `current_df` には、管理者が実際に編集を始めた時点のスナップショット
    (=マトリクス表を描画した際に使った状態)を渡すこと。この関数を呼ぶ直前に
    改めてディスクから最新状態を読み直して渡してはならない。読み直してしまうと、
    管理者が編集している間に他のスタッフが送信した新規申請が「編集前後で消えた
    差分」と誤認識され、取消として上書きされてしまう。
    """
    cur_map = {(r["name"], r["date"]): r["kind"] for _, r in current_df.iterrows()} if not current_df.empty else {}
    new_map = {(r["name"], r["date"]): r["kind"] for _, r in edited_df.iterrows()} if not edited_df.empty else {}

    for key in set(cur_map) | set(new_map):
        old_kind = cur_map.get(key)
        new_kind = new_map.get(key)
        if old_kind == new_kind:
            continue
        name, date = key
        append_kyuka_request(name, date, new_kind if new_kind is not None else CANCELLED_REQUEST_TYPE, path)


def clear_kyuka_log(path: str) -> None:
    """指定月度の希望休・有休ログを削除する(リセット機能用)。"""
    if os.path.exists(path):
        os.remove(path)
    lock_path = _kyuka_lock_path(path)
    if os.path.exists(lock_path):
        os.remove(lock_path)


def kyuka_matrix_date_labels(dates_df: pd.DataFrame) -> list[str]:
    """希望休マトリクス表の列見出し(対象期間の全日付)を返す。"""
    return [f"{d.month}/{d.day}({wd})" for d, wd in zip(dates_df["date"], dates_df["weekday_jp"])]


def build_kyuka_requests_wide(
    requests_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> pd.DataFrame:
    """現在の申請一覧(長形式)を、管理者用マトリクス表(スタッフ×日付)に変換する。

    未申請のセルは店舗別日別シフト表の手動編集と同じ表記(BLANK_LABEL=「（空白）」)
    で埋める(表示・編集プルダウンの選択肢を全社的に統一するため)。
    """
    date_labels = kyuka_matrix_date_labels(dates_df)
    lookup: dict[tuple[str, dt.date], str] = {}
    if not requests_df.empty:
        for _, r in requests_df.iterrows():
            lookup[(r["name"], r["date"])] = r["kind"]

    wide = pd.DataFrame({"スタッフ名": staff_df["name"].tolist()})
    for label, d in zip(date_labels, dates_df["date"]):
        wide[label] = [lookup.get((nm, d), BLANK_LABEL) for nm in staff_df["name"]]
    return wide


def kyuka_requests_wide_to_long(
    wide_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> pd.DataFrame:
    """管理者用マトリクス表(編集後)を、申請一覧(長形式)へ変換する。

    空欄・未知の種別値は無視する(クラッシュしない)。
    """
    date_labels = kyuka_matrix_date_labels(dates_df)
    label_to_date = dict(zip(date_labels, dates_df["date"]))
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))

    rows = []
    for _, r in wide_df.iterrows():
        name = r.get("スタッフ名")
        if name not in name_to_id:
            continue
        for label in date_labels:
            val = str(r.get(label, "")).strip()
            if val not in REQUEST_KINDS:
                continue
            rows.append({"staff_id": name_to_id[name], "name": name, "date": label_to_date[label], "kind": val})
    return pd.DataFrame(rows, columns=["staff_id", "name", "date", "kind"])


# ---------------------------------------------------------------------------
# 最適化結果・手動編集シフトのローカル永続化
# ---------------------------------------------------------------------------

LATEST_SHIFT_PATH = os.path.join(DATA_DIR, "latest_shift_result.csv")
LATEST_SHIFT_META_PATH = os.path.join(DATA_DIR, "latest_shift_meta.json")


def save_shift_result_to_disk(
    shift_df: pd.DataFrame,
    meta: dict,
    shift_path: str = LATEST_SHIFT_PATH,
    meta_path: str = LATEST_SHIFT_META_PATH,
) -> None:
    """最適化結果(または手動編集後の最新シフト)をローカルへ即時保存する。

    最適化の実行完了時・手動編集の確定時・リセット時のいずれからも呼び出すことで、
    ブラウザを閉じたり別ブラウザ/別端末からアクセスした場合でも復元できる。
    """
    os.makedirs(os.path.dirname(shift_path), exist_ok=True)
    out = shift_df.copy()
    if not out.empty:
        out["date"] = out["date"].apply(lambda d: d.isoformat() if isinstance(d, dt.date) else d)
    out.reindex(columns=["date", "store", "staff_id", "name", "emp_type"]).to_csv(
        shift_path, index=False, encoding="utf-8-sig"
    )
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)


def load_shift_result_from_disk(
    shift_path: str = LATEST_SHIFT_PATH,
    meta_path: str = LATEST_SHIFT_META_PATH,
) -> tuple[pd.DataFrame | None, dict]:
    """ローカル保存済みの最新シフト結果とメタ情報を読み込む。

    保存ファイルが無い/壊れている場合は (None, {}) を返す(クラッシュしない)。
    """
    if not os.path.exists(shift_path):
        return None, {}
    try:
        df = pd.read_csv(
            shift_path,
            dtype={"staff_id": str, "name": str, "store": str, "emp_type": str},
            encoding="utf-8-sig",
        )
    except Exception:
        return None, {}
    if not df.empty and "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
        df = df.dropna(subset=["date"])
    df = df.reindex(columns=["date", "store", "staff_id", "name", "emp_type"])

    meta: dict = {}
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
        except Exception:
            meta = {}
    return df, meta


def clear_saved_shift_result(shift_path: str = LATEST_SHIFT_PATH, meta_path: str = LATEST_SHIFT_META_PATH) -> None:
    """ローカル保存済みの最新シフト結果を削除する。"""
    for p in (shift_path, meta_path):
        if os.path.exists(p):
            os.remove(p)


def build_simple_staff_summary(
    shift_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> pd.DataFrame:
    """シフト(長形式)から、簡易版のスタッフ別サマリー(出勤日数・土日出勤・ヘルプ)を作る。

    ローカル保存データから復元したセッションなど、元のソルバー内部変数(嘱託/パートの
    稼働日数レンジ判定など)が無い場合の簡易表示用。
    """
    rows = []
    for row in staff_df.itertuples():
        sid = row.staff_id
        work_days = shift_df[shift_df["staff_id"] == sid] if not shift_df.empty else pd.DataFrame()
        weekend_cnt = int(work_days["date"].apply(lambda d: d.weekday() in (5, 6)).sum()) if not work_days.empty else 0
        help_cnt = 0
        if row.emp_type in ("店長", "正社員", "嘱託") and not work_days.empty:
            help_cnt = int((work_days["store"] != row.home_store).sum())
        rows.append(
            {
                "staff_id": sid,
                "name": row.name,
                "区分": row.emp_type,
                "出勤日数": len(work_days),
                "土日出勤日数": weekend_cnt,
                "ヘルプ出勤日数": help_cnt,
                "勤務日数レンジ(嘱託/パート)": "-",
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 希望休・有休 CSV 一括インポート／テンプレート出力
# ---------------------------------------------------------------------------

_CSV_KIND_LOOKUP = {
    "希望休": "希望休",
    "休": "希望休",
    "公休": "希望休",
    "有休": "有給申請",
    "有給": "有給申請",
    "有給申請": "有給申請",
    "有休申請": "有給申請",  # 旧表記(後方互換)
    "有給確定": "有給申請",  # 廃止区分(後方互換、「有給申請」へ統合)
    "有休確定": "有給申請",  # 旧表記・廃止区分(後方互換、「有給申請」へ統合)
    "絶対休": "絶対休",
}


def _normalize_date_token(s: str) -> str:
    """CSV中の日付表記の揺れ(全角数字・曜日カッコ・年月日表記等)を正規化する。"""
    s = unicodedata.normalize("NFKC", str(s)).strip()
    s = re.sub(r"\(.*?\)", "", s)  # "(水)" 等の曜日カッコを除去
    s = re.sub(r"（.*?）", "", s)
    s = s.replace("年", "/").replace("月", "/")
    s = s.replace("日", "")
    return s.strip()


def _date_label_variants(d: dt.date) -> set[str]:
    """1つの日付が取りうる代表的な文字列表記のバリエーションを列挙する。"""
    variants = set()
    for m in (str(d.month), f"{d.month:02d}"):
        for day in (str(d.day), f"{d.day:02d}"):
            variants.add(f"{m}/{day}")
            variants.add(f"{m}-{day}")
    variants.add(d.isoformat())
    variants.add(f"{d.year}/{d.month}/{d.day}")
    variants.add(f"{d.year}/{d.month:02d}/{d.day:02d}")
    variants.add(f"{d.year}-{d.month:02d}-{d.day:02d}")
    return variants


def _build_date_lookup(dates_df: pd.DataFrame) -> dict[str, dt.date]:
    lookup: dict[str, dt.date] = {}
    for d in dates_df["date"]:
        for v in _date_label_variants(d):
            lookup[v] = d
    return lookup


def generate_requests_csv_template(staff_df: pd.DataFrame, dates_df: pd.DataFrame) -> bytes:
    """全スタッフ×対象期間全日付の、空の希望休入力用CSVテンプレートを生成する。

    日付マトリクス形式(1列目=スタッフ名、2列目以降=各日付)。セルは空欄のまま
    配布し、利用者に「希望休」「有休」「絶対休」のいずれかを記入してもらう。
    """
    date_labels = [f"{d.month}/{d.day}({wd})" for d, wd in zip(dates_df["date"], dates_df["weekday_jp"])]
    template_df = pd.DataFrame({"スタッフ名": staff_df["name"].tolist()})
    for label in date_labels:
        template_df[label] = ""
    return template_df.to_csv(index=False).encode("utf-8-sig")


def parse_requests_csv(
    file_bytes: bytes,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> tuple[pd.DataFrame, list[str]]:
    """希望休・有休CSVを読み込み、requests_df形式(staff_id/name/date/kind)に変換する。

    以下の2形式を自動判別して受け付ける(クラッシュせず、認識できない行/列/値は
    警告メッセージに集約してスキップする):
      - 縦持ち3列形式: スタッフ名, 日付, 種別(希望休/有休/絶対休 等)
      - 日付マトリクス形式: 1列目=スタッフ名、2列目以降=各日付、セル=種別
    """
    warnings: list[str] = []
    empty = pd.DataFrame(columns=["staff_id", "name", "date", "kind"])

    for encoding in ("utf-8-sig", "cp932", "utf-8"):
        try:
            raw = pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False, encoding=encoding)
            break
        except UnicodeDecodeError:
            raw = None
            continue
        except Exception as e:
            return empty, [f"CSVの読み込みに失敗しました: {e}"]
    else:
        return empty, ["CSVの文字コードを認識できませんでした(UTF-8またはShift-JISで保存してください)。"]

    if raw is None or raw.shape[1] < 2:
        return empty, ["CSVに十分な列がありません(最低2列必要です)。"]

    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))
    date_lookup = _build_date_lookup(dates_df)
    rows: list[dict] = []
    col_names = list(raw.columns)
    name_col = col_names[0]

    if raw.shape[1] <= 3:
        # 縦持ち3列形式: スタッフ名, 日付, 種別
        date_col = col_names[1] if len(col_names) > 1 else None
        kind_col = col_names[2] if len(col_names) > 2 else None
        for i, r in raw.iterrows():
            name = str(r[name_col]).strip()
            date_raw = str(r[date_col]).strip() if date_col else ""
            kind_raw = str(r[kind_col]).strip() if kind_col else ""
            if not name and not date_raw and not kind_raw:
                continue
            if name not in name_to_id:
                warnings.append(f"{i + 2}行目: スタッフ名「{name}」が見つかりません。この行はスキップしました。")
                continue
            d = date_lookup.get(_normalize_date_token(date_raw))
            if d is None:
                warnings.append(f"{i + 2}行目: 日付「{date_raw}」を認識できません。この行はスキップしました。")
                continue
            kind = _CSV_KIND_LOOKUP.get(kind_raw.strip())
            if kind is None:
                warnings.append(
                    f"{i + 2}行目: 種別「{kind_raw}」を認識できません"
                    "(希望休/有休/絶対休のいずれかを指定してください)。この行はスキップしました。"
                )
                continue
            rows.append({"staff_id": name_to_id[name], "name": name, "date": d, "kind": kind})
    else:
        # 日付マトリクス形式: 1列目=スタッフ名、2列目以降=各日付列
        resolved_date_cols: dict[str, dt.date] = {}
        for col in col_names[1:]:
            d = date_lookup.get(_normalize_date_token(str(col)))
            if d is None:
                warnings.append(f"列見出し「{col}」を日付として認識できません。この列はスキップしました。")
                continue
            resolved_date_cols[col] = d

        for i, r in raw.iterrows():
            name = str(r[name_col]).strip()
            if not name:
                continue
            if name not in name_to_id:
                warnings.append(f"{i + 2}行目: スタッフ名「{name}」が見つかりません。この行はスキップしました。")
                continue
            for col, d in resolved_date_cols.items():
                val = str(r[col]).strip()
                if not val or val in ("0", "-", "－", "nan", "NaN"):
                    continue
                kind = _CSV_KIND_LOOKUP.get(val)
                if kind is None:
                    warnings.append(
                        f"{name} / {d.month}/{d.day}: 値「{val}」を認識できません"
                        "(希望休/有休/絶対休のいずれかを指定してください)。このセルはスキップしました。"
                    )
                    continue
                rows.append({"staff_id": name_to_id[name], "name": name, "date": d, "kind": kind})

    return pd.DataFrame(rows, columns=["staff_id", "name", "date", "kind"]), warnings


# ---------------------------------------------------------------------------
# 有給休暇「取得可能日・人数枠」の自動算出
# ---------------------------------------------------------------------------

def min_required_employee_bodies_per_day() -> int:
    """1営業日あたり、全7店舗を稼働させるのに最低限必要な「社員(正社員/嘱託)」人数。

    パート併用可能店舗はパートを使う前提（社員1名+パート1名）で最小化、
    徳重店はパターン②（社員1名+パートB+パートC）を使う前提で最小化する。
    """
    general = len(GENERAL_STORES) * 2
    combo = len(COMBO_STORE_PART_ROLES) * 1
    tokushige = 1
    return general + combo + tokushige


def compute_paid_leave_availability(
    dates_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    requests_df: pd.DataFrame,
) -> pd.DataFrame:
    """営業日ごとに「有給取得可能人数枠（社員・嘱託向け）」を自動算出する。

    考え方:
      その日の「社員(店長+正社員+嘱託)」の頭数のうち、既に絶対休/有給申請で
      抜けている人数を除いた「稼働可能人数」から、全店運営に最低限必要な
      人数(min_required_employee_bodies_per_day)を差し引いた余力を、
      その日に新規で有休を割り当てられる人数枠とみなす。
    """
    employee_staff = staff_df[staff_df["emp_type"].isin(EMPLOYEE_TYPES)]
    total_employees = len(employee_staff)
    min_required = min_required_employee_bodies_per_day()

    hard_off_by_date: dict[dt.date, int] = {}
    if not requests_df.empty:
        hard = requests_df[requests_df["kind"].isin(HARD_OFF_KINDS)]
        hard = hard[hard["staff_id"].isin(employee_staff["staff_id"])]
        hard_off_by_date = hard.groupby("date")["staff_id"].nunique().to_dict()

    rows = []
    for _, day in dates_df[dates_df["is_business_day"]].iterrows():
        d = day["date"]
        already_off = hard_off_by_date.get(d, 0)
        available = total_employees - already_off
        surplus = max(0, available - min_required)
        rows.append(
            {
                "date": d,
                "weekday_jp": day["weekday_jp"],
                "day_type": day["day_type"],
                "employee_headcount": total_employees,
                "already_confirmed_off": already_off,
                "min_required_employees": min_required,
                "paid_leave_slots": surplus,
            }
        )
    return pd.DataFrame(rows)


def generate_paid_leave_announcement(
    availability_df: pd.DataFrame,
    period_label: str,
) -> str:
    """スタッフ向けの有給休暇案内テキストをワンクリック生成する。"""
    available_days = availability_df[availability_df["paid_leave_slots"] > 0]

    lines = [
        f"【{period_label}】計画年休のご案内",
        "",
        "下記の日は人員に余力があるため、有給休暇の取得を優先的に受け付けます。",
        "取得をご希望の方は、担当までお申し出ください（先着・調整の上、決定します）。",
        "",
    ]
    if available_days.empty:
        lines.append("※今期間中、余力を確保できる日がありませんでした。個別にご相談ください。")
    else:
        for _, row in available_days.iterrows():
            d: dt.date = row["date"]
            lines.append(
                f"・{d.month}月{d.day}日（{row['weekday_jp']}）"
                f"　取得可能枠：{int(row['paid_leave_slots'])}名"
            )
        lines.append("")
        lines.append(f"合計 {len(available_days)} 日間、取得可能枠あり。")
    lines.append("")
    lines.append("※上記以外の日でも、人員調整が可能な場合は取得できることがあります。")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 最適化結果に基づく「有給休暇 追加取得可能枠」の動的判定
# ---------------------------------------------------------------------------

def _min_required_employees_for_store(
    store: str, part_roles_present: set[str], is_weekend_holiday: bool = False
) -> int:
    """その店舗・その日の実際のパート出勤状況を前提に、成立に最低限必要な社員数を返す。

    optimizer.py の店舗別ハード制約と同じ判定基準に合わせている:
      - 稲沢店・新蟹江店・極楽店: 常に2名(パート不可のため、曜日を問わない)。
      - 大治店: 許可パート(尾澤/野道)のいずれかが出勤していれば1名、なければ2名
        (曜日を問わず同じ判定。野道自身は平日のみ出勤するため、土日祝に
        彼が出勤している状態はそもそも発生しない)。
      - 名古屋中川店・天白植田店: 平日は許可パートが出勤していれば1名、
        なければ2名。土日祝は「社員1名+パート」の2名体制が禁止されているため、
        パートの出勤有無に関わらず常に2名。
      - 徳重店: 平日はパートB・パートC両方が出勤していれば1名、それ以外は2名。
        土日祝は「社員1名+パート2名」の3名体制が禁止されているため、
        パートの出勤有無に関わらず常に2名。
    """
    if store == TOKUSHIGE_STORE:
        if is_weekend_holiday:
            return 2
        return 1 if {"B", "C"} <= part_roles_present else 2
    if store in ("名古屋中川店", "天白植田店"):
        if is_weekend_holiday:
            return 2
        return 1 if part_roles_present & set(COMBO_STORE_PART_ROLES[store]) else 2
    if store == "大治店":
        return 1 if part_roles_present & set(COMBO_STORE_PART_ROLES[store]) else 2
    return 2


_LEAVE_REMOVABLE_TYPES = ("店長", "正社員")  # このロジックで「休みに回す」候補となる区分
# 代替要員(ヘルプ投入)となりうる区分。店長・正社員は全員21日出勤(有給申請者は20日)で
# 固定されており、代替に回すとその分だけ出勤日数が規定を超えてしまうため、
# 代替候補には絶対に含めない(嘱託・パートのみが対象)。
_LEAVE_SUBSTITUTE_TYPES = ("嘱託", "パート")


def compute_post_solve_leave_availability(
    shift_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
    requests_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """最適化【後】の実際のシフト配置から、各営業日の有給追加取得可能枠を算出する。

    代替シミュレーション型のアルゴリズム: 各店舗・各営業日で以下を判定する。
      ケース1(既存の余剰): 現在の在籍人数が成立最低人数を上回っている分は、
        そのまま追加の有休枠となる(店長・正社員が対象、非スキル保有者を優先)。
      ケース2(代替による創出): ちょうど最低人数で足りている場合でも、
        「実際にその日その店舗へ出勤しているスタッフ」の中から店長・正社員を
        1名「休み」に置き換えたと仮定し、その日どこにも出勤しておらず、かつ
        希望休・絶対休・有給申請のいずれも入っていない(=そもそも
        休みたいわけではない)正社員/嘱託の中から
          - この店舗への勤務が許可されている(曜日限定の禁止ルールも含む)
          - 嘱託の場合は月間上限日数にまだ余裕がある
          - 追加しても最大6連勤を超えない
          - スキル要件(抜けた人がスキル保有者なら代替もスキル保有者)を満たす
        代替候補が1名でも見つかれば、その分を有休取得可能枠として計上する。
    いずれの店舗も、既に成立最低人数を下回っている(=不足許容日)場合は対象外とし、
    その日・その店舗の判定のみをスキップする(他の日・他の店舗の計算には一切影響しない)。
    代替候補は同日内で重複して使い回さない。
    休みが確定している(希望休が通った等)スタッフは、有休候補にも代替要員にも
    一切登場しない(実態と矛盾する表示を防ぐ)。
    """
    business_days = dates_df.loc[dates_df["is_business_day"]]

    if shift_df.empty:
        return pd.DataFrame(
            [
                {
                    "date": day["date"],
                    "weekday_jp": day["weekday_jp"],
                    "day_type": day["day_type"],
                    "extra_leave_slots": 0,
                    "detail": "",
                }
                for _, day in business_days.iterrows()
            ]
        )

    staff_emp_type = dict(zip(staff_df["staff_id"], staff_df["emp_type"]))
    staff_has_skill = dict(zip(staff_df["staff_id"], staff_df["has_skill"]))
    staff_home = dict(zip(staff_df["staff_id"], staff_df["home_store"]))
    staff_part_role = dict(zip(staff_df["staff_id"], staff_df["part_role"]))
    staff_name = dict(zip(staff_df["staff_id"], staff_df["name"]))
    staff_max_workdays = dict(zip(staff_df["staff_id"], staff_df.get("max_workdays", pd.Series(dtype=float))))
    staff_allowed = {row.staff_id: set(derive_allowed_stores(row)) for row in staff_df.itertuples()}
    staff_weekend_holiday_forbidden = dict(zip(staff_df["staff_id"], staff_df.get("weekend_holiday_forbidden_stores")))
    staff_sat_sun_forbidden = dict(zip(staff_df["staff_id"], staff_df.get("saturday_sunday_forbidden_stores")))
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))

    # 希望休・絶対休・有給申請のいずれかが入っている(staff_id, date)の集合。
    # このいずれかに該当するスタッフは、そもそも休みたい/休みが確定しているため、
    # 代替要員として「呼び出す」候補には絶対にしない。
    requested_off_pairs: set[tuple[str, dt.date]] = set()
    if requests_df is not None and not requests_df.empty:
        requested_off_pairs = set(zip(requests_df["staff_id"], requests_df["date"]))

    all_days = list(dates_df["date"])
    day_index = {d: i for i, d in enumerate(all_days)}

    work_days_by_staff: dict[str, dict] = {}
    if not shift_df.empty:
        for row in shift_df.itertuples():
            work_days_by_staff.setdefault(row.staff_id, {})[row.date] = row.store
    total_workdays_by_staff = {sid: len(dmap) for sid, dmap in work_days_by_staff.items()}

    def _is_idle(sid: str, d: dt.date) -> bool:
        return d not in work_days_by_staff.get(sid, {})

    def _has_day_capacity(sid: str) -> bool:
        """月間勤務日数レンジの上限に、あと+1日勤務する余力があるかを判定する。"""
        etype = staff_emp_type.get(sid)
        if etype == "嘱託":
            max_d = staff_max_workdays.get(sid)
            if pd.notna(max_d):
                return total_workdays_by_staff.get(sid, 0) < int(max_d)
            return True
        if etype == "パート":
            role = staff_part_role.get(sid)
            _min_d, max_d = PART_ROLE_WORKDAY_RANGE.get(role, (0, len(dates_df)))
            return total_workdays_by_staff.get(sid, 0) < max_d
        return True

    def _six_consecutive_ok(sid: str, d: dt.date) -> bool:
        idx = day_index.get(d)
        if idx is None:
            return True
        worked = work_days_by_staff.get(sid, {})
        for start in range(max(0, idx - 6), idx + 1):
            end = start + 7
            if end > len(all_days):
                continue
            window = all_days[start:end]
            cnt = sum(1 for x in window if x == d or x in worked)
            if cnt > 6:
                return False
        return True

    def _store_allowed_on_date(sid: str, store: str, d: dt.date, is_wh: bool) -> bool:
        if store not in staff_allowed.get(sid, set()):
            return False
        if is_wh and store in (staff_weekend_holiday_forbidden.get(sid) or []):
            return False
        if d.weekday() in (5, 6) and store in (staff_sat_sun_forbidden.get(sid) or []):
            return False
        return True

    rows = []
    for _, day in business_days.iterrows():
        d = day["date"]
        is_wh = bool(day["is_weekend_or_holiday"])
        day_shift = shift_df[shift_df["date"] == d]
        total_slack = 0
        detail_parts = []
        used_substitutes_today: set[str] = set()
        # この日、手動指定でロースターが固定されているスタッフは有休候補から除外する
        # (例: 10/10の中村・山岡・生駒は確定配置のため、その日は休みに回せない)。
        locked_today = {
            name_to_id[n]
            for spec in MANUAL_STORE_ASSIGNMENTS
            if spec["date"] == d
            for n in spec["names"]
            if n in name_to_id
        }

        for store in STORES:
            store_shift = day_shift[day_shift["store"] == store]

            employees_here_all = []  # (sid, has_skill, is_home, etype)
            removable_candidates = []
            part_roles_present: set[str] = set()
            for sid in store_shift["staff_id"]:
                etype = staff_emp_type.get(sid)
                if etype in EMPLOYEE_TYPES:
                    entry = (sid, bool(staff_has_skill.get(sid)), staff_home.get(sid) == store, etype)
                    employees_here_all.append(entry)
                    if etype in _LEAVE_REMOVABLE_TYPES and sid not in locked_today:
                        removable_candidates.append(entry)
                elif etype == "パート":
                    role = staff_part_role.get(sid)
                    if role:
                        part_roles_present.add(role)

            min_required = _min_required_employees_for_store(store, part_roles_present, is_wh)
            current_e = len(employees_here_all)
            if current_e < min_required:
                # 既に不足が許容されている日(手動指定等)は、これ以上減らす判定を行わない。
                continue

            surplus = max(0, current_e - min_required)
            removed_ids: set[str] = set()
            detail_names: list[str] = []

            # --- ケース1: 既存の余剰人員をそのまま有休候補にする ---
            if surplus > 0:
                sorted_candidates = sorted(removable_candidates, key=lambda p: (p[1], p[2]))
                skilled_remaining = sum(1 for e in employees_here_all if e[1])
                for sid, has_skill, _is_home, _etype in sorted_candidates:
                    if len(removed_ids) >= surplus:
                        break
                    if has_skill:
                        if skilled_remaining <= 1:
                            continue
                        skilled_remaining -= 1
                    removed_ids.add(sid)
                detail_names.extend(staff_name.get(s, s) for s in removed_ids)

            # --- ケース2: ちょうど最低人数でも、代替要員を投入できれば1名休める ---
            remaining_for_case2 = [c for c in removable_candidates if c[0] not in removed_ids]
            for sid, has_skill, _is_home, _etype in sorted(remaining_for_case2, key=lambda p: (p[1], p[2])):
                remaining_after_removal = [
                    e for e in employees_here_all if e[0] != sid and e[0] not in removed_ids
                ]
                if len(remaining_after_removal) >= min_required:
                    continue  # 実は代替なしでも足りていた(ケース1で拾いきれなかった分)

                remaining_skilled = sum(1 for e in remaining_after_removal if e[1])
                substitute_id = None
                for cand in staff_df.itertuples():
                    csid = cand.staff_id
                    if csid == sid or csid in removed_ids or csid in used_substitutes_today:
                        continue
                    if staff_emp_type.get(csid) not in _LEAVE_SUBSTITUTE_TYPES:
                        continue
                    if not _store_allowed_on_date(csid, store, d, is_wh):
                        continue
                    if not _is_idle(csid, d):
                        continue
                    if (csid, d) in requested_off_pairs:
                        continue  # 希望休/絶対休/有給申請が入っている人は呼び出さない
                    if not _has_day_capacity(csid):
                        continue
                    if has_skill and remaining_skilled == 0 and not bool(staff_has_skill.get(csid)):
                        continue  # 抜けるのがスキル保有者で、代替もスキルなしでは要件を割る
                    if not _six_consecutive_ok(csid, d):
                        continue
                    substitute_id = csid
                    break

                if substitute_id is not None:
                    removed_ids.add(sid)
                    used_substitutes_today.add(substitute_id)
                    detail_names.append(
                        f"{staff_name.get(sid, sid)}→{staff_name.get(substitute_id, substitute_id)}代替可"
                    )

            if removed_ids:
                total_slack += len(removed_ids)
                detail_parts.append(f"{store}:{'、'.join(detail_names)}")

        rows.append(
            {
                "date": d,
                "weekday_jp": day["weekday_jp"],
                "day_type": day["day_type"],
                "extra_leave_slots": total_slack,
                "detail": " / ".join(detail_parts),
            }
        )
    return pd.DataFrame(rows)


def leave_slot_badge(n: int) -> str:
    """有給追加取得可能枠数をバッジ用の短い文言に変換する。"""
    return f"あと{int(n)}名可能" if n > 0 else "満杯(0名)"


def generate_post_solve_leave_announcement(
    availability_df: pd.DataFrame,
    period_label: str,
) -> str:
    """最適化結果に基づく「有給休暇 追加取得可能日」の社内案内文を生成する。"""
    available_days = availability_df[availability_df["extra_leave_slots"] > 0]

    lines = [
        f"【{period_label}】有給休暇 追加取得可能日のお知らせ",
        "",
        "以下の日程で有給休暇の追加取得が可能です。希望者は申請してください。",
        "",
    ]
    if available_days.empty:
        lines.append("※現在確定しているシフトでは、追加で有給取得可能な日はありません。")
    else:
        for _, row in available_days.iterrows():
            d: dt.date = row["date"]
            lines.append(f"・{d.month}/{d.day}（{row['weekday_jp']}）: {leave_slot_badge(row['extra_leave_slots'])}")
        lines.append("")
        lines.append(f"合計 {len(available_days)} 日間、追加取得可能枠あり。")
    lines.append("")
    lines.append("※先着順・業務都合により調整させていただく場合があります。")
    return "\n".join(lines)


_EMP_TYPE_SORT_ORDER = {"店長": 0, "正社員": 1, "嘱託": 2, "パート": 3}


# ---------------------------------------------------------------------------
# シフト結果の画面上手動編集(玉突き調整)
# ---------------------------------------------------------------------------

BLANK_LABEL = "（空白）"


def manual_shift_date_labels(dates_df: pd.DataFrame) -> list[str]:
    """手動編集テーブルの列見出し(営業日のみ)を返す。"""
    business_days_df = dates_df[dates_df["is_business_day"]]
    return [f"{d.month}/{d.day}({wd})" for d, wd in zip(business_days_df["date"], business_days_df["weekday_jp"])]


def build_manual_shift_wide(
    shift_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> pd.DataFrame:
    """最適化結果(長形式)を、手動編集用のワイド形式(店舗×営業日、1セル1名)に変換する。

    各店舗は STORE_MAX_HEADCOUNT 分の行(枠)を持ち、店舗別日別シフト表と同じ
    レイアウトになる。空きセルは空文字列で表現する。
    """
    business_days_df = dates_df[dates_df["is_business_day"]]
    dates = list(business_days_df["date"])
    date_labels = manual_shift_date_labels(dates_df)

    rows = []
    for store in STORES:
        n_slots = STORE_MAX_HEADCOUNT.get(store, 2)
        per_date_names: dict[dt.date, list[str]] = {}
        for d in dates:
            if shift_df.empty:
                day_rows = pd.DataFrame(columns=["name", "emp_type"])
            else:
                day_rows = shift_df[(shift_df["store"] == store) & (shift_df["date"] == d)]
            names = sorted(
                day_rows["name"].tolist(),
                key=lambda nm: _EMP_TYPE_SORT_ORDER.get(
                    day_rows.loc[day_rows["name"] == nm, "emp_type"].iloc[0], 9
                ),
            )
            per_date_names[d] = names
        for slot in range(n_slots):
            row = {"店舗": store, "枠": slot + 1}
            for d, label in zip(dates, date_labels):
                names = per_date_names[d]
                row[label] = names[slot] if slot < len(names) else ""
            rows.append(row)
    return pd.DataFrame(rows)


def manual_shift_wide_to_long(
    wide_df: pd.DataFrame,
    dates_df: pd.DataFrame,
    staff_df: pd.DataFrame,
) -> pd.DataFrame:
    """手動編集後のワイド形式を、長形式(date/store/staff_id/name/emp_type)へ変換する。

    空欄(空文字列 または BLANK_LABEL)は無視する。存在しない氏名が入っていた
    場合も安全側でスキップする(クラッシュしない)。
    """
    date_labels = manual_shift_date_labels(dates_df)
    label_to_date = dict(zip(date_labels, dates_df.loc[dates_df["is_business_day"], "date"]))
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))
    emp_type_by_name = dict(zip(staff_df["name"], staff_df["emp_type"]))

    rows = []
    for _, r in wide_df.iterrows():
        store = r.get("店舗")
        for label in date_labels:
            name = r.get(label)
            if not name or name == BLANK_LABEL:
                continue
            if name not in name_to_id:
                continue
            rows.append(
                {
                    "date": label_to_date[label],
                    "store": store,
                    "staff_id": name_to_id[name],
                    "name": name,
                    "emp_type": emp_type_by_name[name],
                }
            )
    return pd.DataFrame(rows, columns=["date", "store", "staff_id", "name", "emp_type"])


def _check_store_day_pattern(
    store: str,
    e_count: int,
    part_roles_present: set[str],
    is_weekend_holiday: bool,
) -> tuple[bool, int]:
    """店舗・曜日区分ごとの成立パターンに照らして、現在の在籍構成が有効か判定する。

    optimizer.py の店舗別ハード制約と同じ判定基準(平日/土日祝の違いを含む)。
    戻り値は (成立しているか, 不足人数の目安)。
    """
    if store == TOKUSHIGE_STORE:
        if is_weekend_holiday:
            ok = e_count >= 2
        else:
            ok = e_count >= 2 or (e_count == 1 and {"B", "C"} <= part_roles_present)
    elif store in COMBO_STORE_PART_ROLES and store != "大治店":
        roles = set(COMBO_STORE_PART_ROLES[store])
        if is_weekend_holiday:
            ok = e_count >= 2
        else:
            ok = e_count >= 2 or (e_count == 1 and bool(part_roles_present & roles))
    elif store == "大治店":
        ok = (e_count == 2) or (e_count == 1 and len(part_roles_present) >= 1)
    else:
        ok = e_count == 2
    shortfall = 0 if ok else max(0, 2 - e_count)
    return ok, shortfall


def check_manual_shift_alerts(
    shift_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    dates_df: pd.DataFrame,
) -> dict[str, list[dict]]:
    """手動編集後のシフト(長形式)に対し、重複出勤・人員不足・スキル不在を再判定する。

    戻り値: {"duplicates": [...], "shortages": [...], "skill_issues": [...]}

    shift_df が None/空/想定外の形(必要な列が無い等)の場合は、まだ最適化が
    実行されていない月度などに切り替えた際にクラッシュしないよう、判定を
    行わずに空のアラート(問題なし扱い)を返す。
    """
    alerts: dict[str, list[dict]] = {"duplicates": [], "shortages": [], "skill_issues": []}

    if shift_df is None or not isinstance(shift_df, pd.DataFrame) or shift_df.empty:
        return alerts
    if "date" not in shift_df.columns or "store" not in shift_df.columns or "staff_id" not in shift_df.columns:
        return alerts

    staff_emp_type = dict(zip(staff_df["staff_id"], staff_df["emp_type"]))
    staff_has_skill = dict(zip(staff_df["staff_id"], staff_df["has_skill"]))
    staff_part_role = dict(zip(staff_df["staff_id"], staff_df["part_role"]))
    staff_name = dict(zip(staff_df["staff_id"], staff_df["name"]))

    for (sid, d), grp in shift_df.groupby(["staff_id", "date"]):
        if len(grp) > 1:
            alerts["duplicates"].append(
                {
                    "date": d,
                    "name": staff_name.get(sid, sid),
                    "stores": grp["store"].unique().tolist(),
                }
            )

    business_days_df = dates_df[dates_df["is_business_day"]]
    for _, day in business_days_df.iterrows():
        d = day["date"]
        is_wh = bool(day["is_weekend_or_holiday"])
        day_shift = shift_df[shift_df["date"] == d]
        for store in STORES:
            store_shift = day_shift[day_shift["store"] == store]
            e_count = 0
            part_roles_present: set[str] = set()
            has_skill_here = False
            staffed = False
            for sid in store_shift["staff_id"]:
                etype = staff_emp_type.get(sid)
                if etype in EMPLOYEE_TYPES:
                    e_count += 1
                    staffed = True
                elif etype == "パート":
                    role = staff_part_role.get(sid)
                    if role:
                        part_roles_present.add(role)
                    staffed = True
                if staff_has_skill.get(sid):
                    has_skill_here = True

            ok, shortfall = _check_store_day_pattern(store, e_count, part_roles_present, is_wh)
            if not ok:
                detail = f"社員{e_count}名"
                if part_roles_present:
                    detail += f"+パート{len(part_roles_present)}名"
                alerts["shortages"].append({"date": d, "store": store, "不足数": shortfall, "詳細": detail})
            if staffed and not has_skill_here:
                alerts["skill_issues"].append({"date": d, "store": store})

    return alerts


# ---------------------------------------------------------------------------
# Excelエクスポート
# ---------------------------------------------------------------------------


def _generate_pastel_palette(n: int, lightness: float = 0.84, saturation: float = 0.55) -> list[str]:
    """色相を均等に割り振った、見分けやすいパステル系HEXカラーをn色生成する。

    STAFF_NAME_COLORS に対応がない氏名(=デフォルト19名以外)向けのフォールバック用。
    """
    colors = []
    for i in range(n):
        hue = i / max(n, 1)
        r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
        colors.append(f"{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}")
    return colors


# 全19名の識別カラー(氏名 -> HEX)。隣接しやすいメンバー同士でも一目で見分けられるよう、
# 色相が大きく異なるパステル/識別カラーを個別に選定している。
STAFF_NAME_COLORS: dict[str, str] = {
    "生駒": "BFDFF5",    # パステルブルー(水色)
    "辻本": "C2F0DB",    # ミントグリーン(薄緑)
    "内田": "FFF3B8",    # ソフトイエロー(薄黄)
    "小林": "FFC9CE",    # コーラルピンク(薄紅)
    "加藤": "E1D3F5",    # ラベンダー(薄紫)
    "田中": "FFDCB8",    # アプリコット(薄橙)
    "長瀬": "A9E8E8",    # スカイシアン(明るい青緑)
    "真田": "E4F0A0",    # レモンライム(黄緑)
    "若林": "F7C9DC",    # ローズピンク(薄ピンク)
    "中村": "A0E4DC",    # ターコイズ(青緑)
    "若松": "E9DCBB",    # サンドベージュ(薄茶/ベージュ)
    "吉田": "E2D48A",    # オリーブゴールド(薄オリーブ)
    "竹内": "D6B8E6",    # モーヴ(薄藤色)
    "山岡": "DCDCDC",    # ペールグレー/シルバー(明灰色)
    "尾澤": "FFB6A3",    # サーモン(薄サーモンピンク)
    "不動野": "A0E0BC",  # エメラルド(薄深緑)
    "前田": "E3AEDD",    # オーキッド(赤紫)
    "柴田": "DCF3FA",    # アイスブルー(ごく薄い青)
    "野道": "FFDAC0",    # ピーチ(薄桃色)
}


def assign_staff_colors(staff_df: pd.DataFrame) -> dict[str, str]:
    """スタッフ1名ごとに固有の識別背景色(HEX下6桁)を割り当てる。

    まず STAFF_NAME_COLORS の氏名一致を優先し、該当がないスタッフ(氏名変更時や
    デフォルト以外の追加スタッフ)には自動生成したパステルカラーを割り当てる。
    """
    fallback_palette = _generate_pastel_palette(len(staff_df))
    colors: dict[str, str] = {}
    for i, row in enumerate(staff_df.itertuples()):
        colors[row.staff_id] = STAFF_NAME_COLORS.get(row.name, fallback_palette[i % len(fallback_palette)])
    return colors


def build_export_workbook(
    shift_df: pd.DataFrame,
    dates_df: pd.DataFrame,
    staff_df: pd.DataFrame,
    requests_df: pd.DataFrame | None = None,
) -> bytes:
    """店舗別日別シフト表・スタッフ別出勤一覧表の2シート構成Excelを生成する。

    shift_df: columns = [date, store, staff_id, name, emp_type]  (出勤が確定した行のみ)

    「店舗別日別シフト表」は現場配布・掲示にそのまま使えるよう、1セルに1名のみを
    配置する縦分割レイアウトとし、下段に休日スタッフ一覧(赤文字)を付す。
    休日スタッフのうち、希望休・絶対休・有給申請のいずれかを本人が
    申請していた場合は赤文字＋太字とし、AIが自動割り当てした公休(通常の赤文字)
    と一目で区別できるようにする。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dates = list(dates_df["date"])
    date_labels = [
        f"{d.month}/{d.day}({row.weekday_jp})" + ("\n休業" if getattr(row, "is_special_closure", False) else "")
        for d, row in zip(dates, dates_df.itertuples())
    ]
    n_date_cols = len(dates)
    total_cols = 1 + n_date_cols  # 列A=店舗名/見出し、以降は日付列

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "店舗別日別シフト表"

    # 1行目(ヘッダー行)のフォントサイズ。太字・中央揃えは従来通り維持する。
    HEADER_FONT_SIZE = 9

    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="404040")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True, size=HEADER_FONT_SIZE)
    store_fills = [PatternFill("solid", fgColor="EDEDED"), PatternFill("solid", fgColor="FFFFFF")]
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    red_font = Font(color="FF0000")
    red_bold_font = Font(color="FF0000", bold=True)
    requested_off_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 希望休/絶対休/有給申請のいずれかを本人が申請していた(氏名,日付)の集合。
    # 該当する休日スタッフ名は赤文字＋太字にして、AI自動割り当ての公休と区別する。
    requested_off_pairs: set[tuple[str, dt.date]] = set()
    if requests_df is not None and not requests_df.empty:
        requested_off_pairs = set(zip(requests_df["name"], requests_df["date"]))

    # 日付見出し(土曜/日曜・祝日/平日/特別休業日)の強調配色
    SAT_FILL, SAT_FONT = "DCE6F1", "002060"
    SUN_HOLIDAY_FILL, SUN_HOLIDAY_FONT = "FCE4D6", "C00000"
    WEEKDAY_HEADER_FILL, WEEKDAY_HEADER_FONT = "E2EFDA", "375623"
    SPECIAL_CLOSURE_FILL, SPECIAL_CLOSURE_FONT = "BFBFBF", "404040"

    special_closure_by_date = dict(zip(dates_df["date"], dates_df.get("is_special_closure", False)))

    def _date_header_style(d: dt.date) -> tuple[str, str]:
        if special_closure_by_date.get(d):
            return SPECIAL_CLOSURE_FILL, SPECIAL_CLOSURE_FONT
        if jpholiday.is_holiday(d) or d.weekday() == 6:
            return SUN_HOLIDAY_FILL, SUN_HOLIDAY_FONT
        if d.weekday() == 5:
            return SAT_FILL, SAT_FONT
        return WEEKDAY_HEADER_FILL, WEEKDAY_HEADER_FONT

    def _set_bottom_border(row_idx: int, style: Side):
        for col in range(1, total_cols + 1):
            c = ws1.cell(row=row_idx, column=col)
            b = c.border
            c.border = Border(left=b.left, right=b.right, top=b.top, bottom=style)

    # --- ヘッダー行 ---------------------------------------------------------
    ws1.cell(row=1, column=1, value="店舗").font = header_font
    ws1.cell(row=1, column=1).fill = header_fill
    ws1.cell(row=1, column=1).alignment = center
    ws1.cell(row=1, column=1).border = cell_border
    for j, (label, d) in enumerate(zip(date_labels, dates), start=2):
        fill_hex, font_hex = _date_header_style(d)
        c = ws1.cell(row=1, column=j, value=label)
        c.font = Font(color=font_hex, bold=True, size=HEADER_FONT_SIZE)
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.alignment = center
        c.border = cell_border

    # --- スタッフ固有のパステル背景色 -----------------------------------------
    staff_colors = assign_staff_colors(staff_df)
    name_to_id = dict(zip(staff_df["name"], staff_df["staff_id"]))

    def _staff_fill(name: str) -> PatternFill | None:
        sid = name_to_id.get(name)
        color = staff_colors.get(sid)
        return PatternFill("solid", fgColor=color) if color else None

    # --- 店舗ごとの出勤者ブロック(1セル1名・縦分割) ----------------------------
    current_row = 2
    emp_type_by_staff = dict(zip(shift_df["staff_id"], shift_df["emp_type"])) if not shift_df.empty else {}

    # 「店舗別日別シフト表」シートは、実際の店舗別上限人数(STORE_MAX_HEADCOUNT、
    # 2名または3名)に関わらず、全7店舗を一律「縦4行(1〜4枠)」で出力する
    # (印刷・現場配布時のレイアウトを店舗間で統一するための表示上の仕様であり、
    # ソルバー側の実際の店舗別人員上限キャップ(2名/3名)を変更するものではない)。
    # 実人数が4名に満たない店舗・日は、余った枠が自動的に空欄になる。
    EXCEL_STORE_SLOT_ROWS = 4

    for idx, store in enumerate(STORES):
        n_rows = EXCEL_STORE_SLOT_ROWS
        start_row = current_row
        fill = store_fills[idx % 2]

        ws1.merge_cells(start_row=start_row, start_column=1, end_row=start_row + n_rows - 1, end_column=1)
        store_cell = ws1.cell(row=start_row, column=1, value=store)
        store_cell.font = Font(bold=True)
        store_cell.alignment = center
        store_cell.fill = fill
        store_cell.border = cell_border

        for j, d in enumerate(dates, start=2):
            if special_closure_by_date.get(d):
                # 特別休業日: 全店休業のため、店舗行はグレーアウトし「休業」と表示する。
                for r_offset in range(n_rows):
                    row_idx = start_row + r_offset
                    cell = ws1.cell(row=row_idx, column=j, value="休業")
                    cell.alignment = center
                    cell.font = Font(color=SPECIAL_CLOSURE_FONT)
                    cell.fill = PatternFill("solid", fgColor=SPECIAL_CLOSURE_FILL)
                    cell.border = cell_border
                continue

            if shift_df.empty:
                day_rows = pd.DataFrame(columns=["name", "emp_type"])
            else:
                day_rows = shift_df[(shift_df["store"] == store) & (shift_df["date"] == d)]
            names = sorted(
                day_rows["name"].tolist(),
                key=lambda nm: _EMP_TYPE_SORT_ORDER.get(
                    day_rows.loc[day_rows["name"] == nm, "emp_type"].iloc[0], 9
                ),
            )
            for r_offset in range(n_rows):
                row_idx = start_row + r_offset
                value = names[r_offset] if r_offset < len(names) else ""
                cell = ws1.cell(row=row_idx, column=j, value=value)
                cell.alignment = center
                cell.fill = _staff_fill(value) or fill
                cell.border = cell_border

        for r_offset in range(n_rows):
            ws1.cell(row=start_row + r_offset, column=1).fill = fill

        _set_bottom_border(start_row + n_rows - 1, thick)
        current_row += n_rows

    # --- 休日スタッフ一覧セクション(赤文字) ------------------------------------
    current_row += 1
    section_row = current_row
    ws1.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=total_cols)
    section_cell = ws1.cell(row=section_row, column=1, value="【休日スタッフ（公休・有休・希望休）】")
    section_cell.font = Font(bold=True, color="C00000")
    section_cell.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1

    all_staff_names = list(staff_df["name"])
    off_lists: list[list[str]] = []
    for d in dates:
        working_names = set(shift_df[shift_df["date"] == d]["name"]) if not shift_df.empty else set()
        off_lists.append([n for n in all_staff_names if n not in working_names])
    max_off = max((len(lst) for lst in off_lists), default=0)
    max_off = max(max_off, 1)

    off_start_row = current_row
    ws1.merge_cells(start_row=off_start_row, start_column=1, end_row=off_start_row + max_off - 1, end_column=1)
    off_label_cell = ws1.cell(row=off_start_row, column=1, value="休日スタッフ")
    off_label_cell.font = Font(bold=True)
    off_label_cell.alignment = center
    off_label_cell.border = cell_border

    for col_idx, off_names in enumerate(off_lists):
        j = col_idx + 2
        d = dates[col_idx]
        for r_offset in range(max_off):
            row_idx = off_start_row + r_offset
            value = off_names[r_offset] if r_offset < len(off_names) else ""
            cell = ws1.cell(row=row_idx, column=j, value=value)
            cell.alignment = center
            cell.border = cell_border
            is_requested = (value, d) in requested_off_pairs
            if is_requested:
                # 事前申請休みは、個人固有色より優先して黄色背景＋赤太字にする。
                cell.font = red_bold_font
                cell.fill = requested_off_fill
            else:
                cell.font = red_font
                staff_fill = _staff_fill(value)
                if staff_fill:
                    cell.fill = staff_fill

    for r_offset in range(max_off):
        ws1.cell(row=off_start_row + r_offset, column=1).border = cell_border

    # --- 列幅・行高・印刷設定 ---------------------------------------------------
    ws1.column_dimensions[get_column_letter(1)].width = 16
    # B列(2列目)〜AG列(33列目)は日別の各列として一律 43px相当(幅5.4)に統一する。
    # 実データが33列に満たない月でも、テンプレートとしてAG列まで幅を揃えておく。
    for j in range(2, 34):
        ws1.column_dimensions[get_column_letter(j)].width = 5.4
    for j in range(34, total_cols + 1):
        ws1.column_dimensions[get_column_letter(j)].width = 13

    # 1行目(ヘッダー)は 52px相当(39pt)、2行目〜店舗ブロック最終行(2行目+7店舗×
    # 4枠-1=29行目)は 38px相当(28.5pt)に統一する(店舗ブロックの行数を固定値で
    # 再計算せず、実際に描画したEXCEL_STORE_SLOT_ROWS×STORES数から動的に導出する)。
    header_row_end = 1
    store_rows_end = header_row_end + EXCEL_STORE_SLOT_ROWS * len(STORES)  # 2行目+28行=29行目
    ws1.row_dimensions[1].height = 39.0
    for r in range(2, store_rows_end + 1):
        ws1.row_dimensions[r].height = 28.5

    ws1.freeze_panes = "B2"
    ws1.page_setup.orientation = "landscape"
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A3
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.print_title_rows = "1:1"

    # --- シート2: スタッフ別出勤一覧表 -----------------------------------
    ws2 = wb.create_sheet("スタッフ別出勤一覧表")
    header2 = ["スタッフ", "区分"] + date_labels + ["出勤日数"]
    for j, label in enumerate(header2, start=1):
        c = ws2.cell(row=1, column=j, value=label)
        c.alignment = center
        c.border = cell_border
        if 3 <= j <= 2 + n_date_cols:
            fill_hex, font_hex = _date_header_style(dates[j - 3])
            c.font = Font(color=font_hex, bold=True, size=HEADER_FONT_SIZE)
            c.fill = PatternFill("solid", fgColor=fill_hex)
        else:
            c.font = header_font
            c.fill = header_fill

    for i, s in enumerate(staff_df.itertuples(), start=2):
        sid = s.staff_id
        row_fill = PatternFill("solid", fgColor=staff_colors[sid]) if sid in staff_colors else None
        name_cell = ws2.cell(row=i, column=1, value=s.name)
        name_cell.alignment = center
        name_cell.border = cell_border
        emp_cell = ws2.cell(row=i, column=2, value=s.emp_type)
        emp_cell.alignment = center
        emp_cell.border = cell_border
        if row_fill:
            name_cell.fill = row_fill
            emp_cell.fill = row_fill
        work_count = 0
        for j, d in enumerate(dates, start=3):
            match = shift_df[(shift_df["staff_id"] == sid) & (shift_df["date"] == d)] if not shift_df.empty else pd.DataFrame()
            value = match.iloc[0]["store"] if not match.empty else ""
            if not match.empty:
                work_count += 1
            day_cell = ws2.cell(row=i, column=j, value=value)
            day_cell.alignment = center
            day_cell.border = cell_border
            if row_fill:
                day_cell.fill = row_fill
        count_cell = ws2.cell(row=i, column=2 + n_date_cols + 1, value=work_count)
        count_cell.alignment = center
        count_cell.border = cell_border
        if row_fill:
            count_cell.fill = row_fill

    ws2.column_dimensions[get_column_letter(1)].width = 12
    ws2.column_dimensions[get_column_letter(2)].width = 8
    # 日別の各列(C列以降、店舗別日別シフト表のB〜AG列に相当)は同じく幅5.4に統一する。
    for j in range(3, 3 + n_date_cols):
        ws2.column_dimensions[get_column_letter(j)].width = 5.4
    ws2.column_dimensions[get_column_letter(3 + n_date_cols)].width = 10

    # 1行目(ヘッダー)〜18行目の行高を 52px相当(39pt)に統一する。
    for r in range(1, 19):
        ws2.row_dimensions[r].height = 39.0

    ws2.freeze_panes = "C2"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A3
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
